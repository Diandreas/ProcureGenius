from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Q, Sum, Count, F, Value
from django.db.models.functions import Greatest, Coalesce
from django.db import models
from django.utils import timezone
from datetime import timedelta

from apps.suppliers.models import Supplier, SupplierCategory, SupplierProduct
from apps.purchase_orders.models import PurchaseOrder, PurchaseOrderItem
from apps.invoicing.models import Invoice, InvoiceItem, Product, ProductCategory, Warehouse
from apps.accounts.models import Client
from apps.core.permissions import HasModuleAccess
from apps.core.modules import Modules, user_has_module_access
from apps.core.organization_mixin import OrganizationFilterMixin, OrganizationClientFilterMixin

from .serializers import (
    SupplierSerializer, SupplierCategorySerializer, SupplierProductSerializer,
    PurchaseOrderSerializer, PurchaseOrderItemSerializer,
    InvoiceSerializer, InvoiceItemSerializer,
    ProductSerializer, ClientSerializer,
    ProductCategorySerializer, WarehouseSerializer,
    DashboardStatsSerializer
)


class SupplierCategoryViewSet(viewsets.ModelViewSet):
    """ViewSet pour les catégories de fournisseurs"""
    queryset = SupplierCategory.objects.all()
    serializer_class = SupplierCategorySerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    required_module = Modules.SUPPLIERS
    # NOTE: SupplierCategory is a global model shared across organizations
    # No organization filter needed


class SupplierViewSet(OrganizationFilterMixin, viewsets.ModelViewSet):
    """ViewSet pour les fournisseurs"""
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    required_module = Modules.SUPPLIERS
    organization_field = 'organization'
    filterset_fields = ['status', 'province', 'is_local', 'is_active']
    search_fields = ['name', 'contact_person', 'email', 'city']
    ordering_fields = ['name', 'rating', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        # First apply organization filter
        queryset = super().get_queryset()
        
        # Filtre par catégorie
        category_id = self.request.query_params.get('category')
        if category_id:
            queryset = queryset.filter(categories__id=category_id)
        
        # Filtre par note minimale
        min_rating = self.request.query_params.get('min_rating')
        if min_rating:
            queryset = queryset.filter(rating__gte=min_rating)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def toggle_status(self, request, pk=None):
        """Active/Désactive un fournisseur"""
        supplier = self.get_object()
        if supplier.status == 'active':
            supplier.status = 'inactive'
        else:
            supplier.status = 'active'
        supplier.save()
        serializer = self.get_serializer(supplier)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Statistiques détaillées du fournisseur"""
        supplier = self.get_object()

        # Bons de commande du fournisseur
        purchase_orders = PurchaseOrder.objects.filter(supplier=supplier)

        # Statistiques financières
        total_spent = purchase_orders.aggregate(
            total=Sum('total_amount')
        )['total'] or 0

        # Statistiques par statut
        po_stats = purchase_orders.values('status').annotate(
            count=Count('id'),
            total_amount=Sum('total_amount')
        )

        # Produits/services les plus achetés
        top_products = PurchaseOrderItem.objects.filter(
            purchase_order__supplier=supplier
        ).values('description', 'product_reference').annotate(
            total_quantity=Sum('quantity'),
            total_value=Sum('total_price'),
            order_count=Count('purchase_order', distinct=True)
        ).order_by('-total_value')[:10]

        # Activité récente (6 derniers mois)
        from datetime import datetime, timedelta
        six_months_ago = datetime.now() - timedelta(days=180)
        recent_orders = purchase_orders.filter(
            created_at__gte=six_months_ago
        ).values('created_at__month', 'created_at__year').annotate(
            count=Count('id'),
            total_amount=Sum('total_amount')
        ).order_by('created_at__year', 'created_at__month')

        # Performance metrics
        avg_delivery_time = None  # À implémenter si on a des données de livraison
        on_time_delivery_rate = None  # À implémenter

        return Response({
            'supplier_id': str(supplier.id),
            'supplier_name': supplier.name,
            'financial_stats': {
                'total_spent': float(total_spent),
                'total_orders': purchase_orders.count(),
                'average_order_value': float(total_spent / purchase_orders.count()) if purchase_orders.count() > 0 else 0,
            },
            'purchase_orders': {
                'recent': list(purchase_orders.order_by('-created_at')[:5].values(
                    'id', 'po_number', 'title', 'status', 'total_amount', 'created_at'
                )),
                'by_status': list(po_stats),
                'total_count': purchase_orders.count(),
            },
            'top_products': list(top_products),
            'activity_timeline': list(recent_orders),
            'performance_metrics': {
                'avg_delivery_time': avg_delivery_time,
                'on_time_delivery_rate': on_time_delivery_rate,
                'total_suppliers_rank': None,  # À implémenter
            }
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export des fournisseurs en CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="suppliers.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Contact', 'Email', 'Phone', 'City', 'Province', 'Status', 'Rating'])
        
        for supplier in self.get_queryset():
            try:
                status_display = supplier.get_status_display() if hasattr(supplier, 'get_status_display') else supplier.status
            except:
                status_display = supplier.status or ''
            
            writer.writerow([
                supplier.name or '',
                supplier.contact_person or '',
                supplier.email or '',
                supplier.phone or '',
                supplier.city or '',
                supplier.province or '',
                status_display,
                str(supplier.rating) if supplier.rating else '0'
            ])
        
        return response
    
    @action(detail=True, methods=['get'], url_path='pdf-report')
    def generate_pdf_report(self, request, pk=None):
        """Générer un rapport PDF pour un fournisseur"""
        from django.http import HttpResponse
        import traceback
        
        try:
            supplier = self.get_object()
            
            # Générer le PDF avec WeasyPrint
            from .services.report_generator_weasy import generate_supplier_report_pdf
            pdf_buffer = generate_supplier_report_pdf(supplier, request.user)
            
            # Créer la réponse HTTP
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            
            supplier_name = getattr(supplier, 'name', 'fournisseur') or 'fournisseur'
            filename = f"rapport-fournisseur-{supplier_name.replace(' ', '_')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(response.content)
            
            return response
            
        except ImportError as e:
            print(f"ImportError génération PDF fournisseur: {e}")
            traceback.print_exc()
            return Response(
                {'error': 'Service de génération PDF non disponible', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            print(f"Erreur génération PDF fournisseur: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du PDF: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class SupplierProductViewSet(OrganizationFilterMixin, viewsets.ModelViewSet):
    """ViewSet pour les relations Fournisseur-Produit"""
    queryset = SupplierProduct.objects.all()
    serializer_class = SupplierProductSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    required_module = Modules.SUPPLIERS
    organization_field = 'supplier__organization'
    filterset_fields = ['supplier', 'product', 'is_preferred', 'is_active']
    search_fields = ['supplier__name', 'product__name', 'supplier_reference']
    ordering_fields = ['created_at', 'supplier_price', 'lead_time_days']
    ordering = ['-is_preferred', 'supplier__name']

    def get_queryset(self):
        # First apply organization filter
        queryset = super().get_queryset()

        # Filtre par fournisseur
        supplier_id = self.request.query_params.get('supplier_id')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)

        # Filtre par produit
        product_id = self.request.query_params.get('product_id')
        if product_id:
            queryset = queryset.filter(product_id=product_id)

        return queryset.select_related('supplier', 'product')

    @action(detail=False, methods=['get'])
    def by_supplier(self, request):
        """Liste des produits d'un fournisseur spécifique"""
        supplier_id = request.query_params.get('supplier_id')
        if not supplier_id:
            return Response(
                {'error': 'supplier_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        supplier_products = self.get_queryset().filter(supplier_id=supplier_id)
        serializer = self.get_serializer(supplier_products, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_product(self, request):
        """Liste des fournisseurs d'un produit spécifique"""
        product_id = request.query_params.get('product_id')
        if not product_id:
            return Response(
                {'error': 'product_id parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        product_suppliers = self.get_queryset().filter(product_id=product_id)
        serializer = self.get_serializer(product_suppliers, many=True)
        return Response(serializer.data)


class ProductViewSet(OrganizationFilterMixin, viewsets.ModelViewSet):
    """ViewSet pour les produits"""
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticated]
    # required_module removed - products/services should be accessible regardless of module activation
    organization_field = 'organization'  # Product has organization FK
    filterset_fields = ['is_active', 'product_type']
    search_fields = ['name', 'reference', 'description']
    ordering_fields = ['name', 'price', 'stock_quantity']
    ordering = ['name']

    def get_queryset(self):
        # First apply organization filter
        queryset = super().get_queryset()
        

        # Filtre par fournisseur (via la relation SupplierProduct)
        supplier_id = self.request.query_params.get('supplier')
        if supplier_id:
            queryset = queryset.filter(supplier_products__supplier_id=supplier_id).distinct()

        # Filtre par statut de stock (inclut stock des lots disponibles/ouverts)
        stock_status = self.request.query_params.get('stock_status')
        if stock_status:
            queryset = queryset.filter(product_type='physical').annotate(
                _batch_stock=Coalesce(
                    Sum('batches__quantity_remaining',
                        filter=Q(batches__status__in=['available', 'opened'])),
                    Value(0)
                ),
                _effective_stock=Greatest(F('stock_quantity'), F('_batch_stock'))
            )
            if stock_status == 'out_of_stock':
                queryset = queryset.filter(_effective_stock=0)
            elif stock_status == 'low_stock':
                queryset = queryset.filter(
                    _effective_stock__gt=0,
                    _effective_stock__lte=F('low_stock_threshold')
                )
            elif stock_status == 'ok':
                queryset = queryset.filter(
                    _effective_stock__gt=F('low_stock_threshold')
                )
        
        # Exclure les consommables labo UNIQUEMENT sur le listing
        # (list, export_csv, low_stock…) — jamais sur retrieve/update/destroy
        # pour éviter que des produits soient inaccessibles après création.
        list_only_actions = {'list', 'low_stock', 'export_csv', 'expired_report',
                             'batch_stats', 'by_supplier', 'by_product'}
        action = getattr(self, 'action', None)
        if action in list_only_actions:
            try:
                queryset = queryset.exclude(
                    Q(category__slug__in=[
                        'lab_consumables', 'lab-consumables',
                        'laboratory', 'lab-tests', 'lab_tests'
                    ]) |
                    Q(linked_lab_tests__isnull=False)
                )
            except Exception:
                pass

        return queryset.distinct()

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Produits avec stock faible"""
        products = self.get_queryset().filter(
            product_type='physical',
            stock_quantity__lte=F('low_stock_threshold')
        )
        serializer = self.get_serializer(products, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export inventaire complet : 5 onglets + graphiques manager"""
        import io
        from collections import defaultdict
        from decimal import Decimal
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone
        from django.db.models import Sum, Count, Q

        today = timezone.now().date()
        org = getattr(request.user, 'organization', None)

        # ═══════════════════════════════════════════════════════════════════════
        # PALETTE & HELPERS
        # ═══════════════════════════════════════════════════════════════════════
        C = {
            'navy':      '1E3A5F',
            'blue':      '2980B9',
            'blue_lt':   'AED6F1',
            'teal':      '1ABC9C',
            'green':     '27AE60',
            'green_lt':  'A9DFBF',
            'orange':    'E67E22',
            'orange_lt': 'FAD7A0',
            'red':       'C0392B',
            'red_lt':    'F1948A',
            'purple':    '8E44AD',
            'gold':      'F1C40F',
            'grey_hdr':  'ECF0F1',
            'grey_row':  'F8FAFB',
            'white':     'FFFFFF',
            'text':      '2C3E50',
            'text_lt':   '7F8C8D',
        }

        PRICE_FMT  = '#,##0'
        PRICE2_FMT = '#,##0.00'
        PCT_FMT    = '0.0"%"'
        INT_FMT    = '#,##0'

        def _side(color='D5D8DC'):
            return Side(style='thin', color=color)

        def _border(color='D5D8DC'):
            s = _side(color)
            return Border(left=s, right=s, top=s, bottom=s)

        def _hdr(cell, text, bg=None, fg=None, size=10, wrap=False):
            cell.value = text
            cell.fill = PatternFill(start_color=bg or C['navy'],
                                    end_color=bg or C['navy'], fill_type='solid')
            cell.font = Font(bold=True, color=fg or C['white'], size=size)
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=wrap)
            cell.border = _border('BDC3C7')

        def _cell(cell, value, row_i, fmt=None, bold=False, fg=None,
                  bg=None, align='center', wrap=False):
            cell.value = value
            _bg = bg or (C['grey_row'] if row_i % 2 == 0 else C['white'])
            cell.fill = PatternFill(start_color=_bg, end_color=_bg, fill_type='solid')
            cell.font = Font(bold=bold, color=fg or C['text'], size=10)
            cell.alignment = Alignment(horizontal=align, vertical='center',
                                       wrap_text=wrap)
            cell.border = _border()
            if fmt:
                cell.number_format = fmt

        def _kpi_block(ws, row, col, label, value, bg, fg=None):
            """Bloc KPI 2 lignes : label (petit) + valeur (grand)"""
            lc = ws.cell(row, col, label)
            lc.font = Font(bold=False, size=9, color=fg or C['white'])
            lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            lc.alignment = Alignment(horizontal='center', vertical='bottom')

            vc = ws.cell(row + 1, col, value)
            vc.font = Font(bold=True, size=14, color=fg or C['white'])
            vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            vc.alignment = Alignment(horizontal='center', vertical='top')

        def _title_row(ws, row, text, col_span_end, bg=None, size=12):
            c = ws.cell(row, 1, text)
            c.font = Font(bold=True, size=size, color=C['white'])
            c.fill = PatternFill(start_color=bg or C['navy'],
                                 end_color=bg or C['navy'], fill_type='solid')
            c.alignment = Alignment(horizontal='left', vertical='center',
                                    indent=1)
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=col_span_end)
            ws.row_dimensions[row].height = 24

        def _section_hdr(ws, row, text, col_end, bg=None):
            c = ws.cell(row, 1, text)
            c.font = Font(bold=True, size=10, color=C['white'])
            c.fill = PatternFill(start_color=bg or C['blue'],
                                 end_color=bg or C['blue'], fill_type='solid')
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=col_end)
            ws.row_dimensions[row].height = 18

        def _add_chart(ws_charts, chart, anchor, title, title_size=11):
            chart.title = title
            chart.style  = 2
            chart.title  = title
            ws_charts.add_chart(chart, anchor)

        # ═══════════════════════════════════════════════════════════════════════
        # DONNÉES SOURCES
        # ═══════════════════════════════════════════════════════════════════════
        from apps.invoicing.models import ProductBatch, StockMovement, InvoiceItem

        all_products = list(
            self.get_queryset()
            .select_related('category', 'supplier')
            .prefetch_related('batches')
        )
        physical = [p for p in all_products if p.product_type == 'physical']
        services = [p for p in all_products if p.product_type != 'physical']

        def _stock(product):
            """
            Calcule le stock réel en utilisant les lots déjà préchargés (pas de requête DB).
            Utilise product.batches.all() — compatible avec prefetch_related('batches').
            Retourne max(stock_quantity, somme lots available/opened).
            """
            batch_total = sum(
                (b.quantity_remaining or 0)
                for b in product.batches.all()
                if b.status in ('available', 'opened')
            )
            return max(product.stock_quantity or 0, batch_total)

        # Ventes 30 derniers jours par produit
        since_30 = today - __import__('datetime').timedelta(days=30)
        sales_30 = (
            InvoiceItem.objects
            .filter(
                invoice__status__in=['paid', 'sent', 'overdue'],
                invoice__created_at__date__gte=since_30,
                product__isnull=False,
            )
            .values('product_id')
            .annotate(qty=Sum('quantity'), revenue=Sum('total_price'))
        )
        sales_map = {str(r['product_id']): r for r in sales_30}

        # Pertes 30j
        losses_30 = (
            StockMovement.objects
            .filter(movement_type='loss', created_at__date__gte=since_30)
            .values('product_id')
            .annotate(qty=Sum('quantity'), val=Sum('loss_value'))
        )
        loss_map = {str(r['product_id']): r for r in losses_30}

        # Pré-calcul du stock réel pour chaque produit physique (1 seul passage)
        stock_cache = {str(p.id): _stock(p) for p in physical}

        # Stats globales
        total_stock_value   = sum(stock_cache[str(p.id)] * float(p.cost_price or 0) for p in physical)
        total_retail_value  = sum(stock_cache[str(p.id)] * float(p.price or 0)      for p in physical)
        total_potential_margin = total_retail_value - total_stock_value

        ok_count  = sum(1 for p in physical if stock_cache[str(p.id)] > (p.low_stock_threshold or 0))
        low_count = sum(1 for p in physical if 0 < stock_cache[str(p.id)] <= (p.low_stock_threshold or 0))
        out_count = sum(1 for p in physical if stock_cache[str(p.id)] <= 0)

        expiring_30 = sum(
            1 for p in physical
            if p.expiration_date and p.expiration_date <= today + __import__('datetime').timedelta(days=30)
        )
        batch_expiring = 0
        try:
            batch_expiring = ProductBatch.objects.filter(
                product__in=physical,
                expiry_date__lte=today + __import__('datetime').timedelta(days=30),
                expiry_date__gte=today,
                quantity_remaining__gt=0,
            ).count()
        except Exception:
            pass

        total_loss_value = sum(float(r.get('val') or 0) for r in loss_map.values())
        total_revenue_30 = sum(float(r.get('revenue') or 0) for r in sales_map.values())

        # ═══════════════════════════════════════════════════════════════════════
        # WORKBOOK
        # ═══════════════════════════════════════════════════════════════════════
        wb = openpyxl.Workbook()

        # ─────────────────────────────────────────────────────────────────────
        # ONGLET 1 : RÉSUMÉ EXÉCUTIF (Dashboard KPI)
        # ─────────────────────────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Résumé Exécutif"
        ws_sum.sheet_view.showGridLines = False
        ws_sum.column_dimensions['A'].width = 3
        for col in ['B','C','D','E','F','G','H','I','J','K','L','M','N']:
            ws_sum.column_dimensions[col].width = 14

        # Bandeau titre
        ws_sum.merge_cells('B2:N2')
        title_c = ws_sum['B2']
        title_c.value = f"RAPPORT INVENTAIRE  —  {today.strftime('%d/%m/%Y')}"
        title_c.font = Font(bold=True, size=18, color=C['white'])
        title_c.fill = PatternFill(start_color=C['navy'], end_color=C['navy'], fill_type='solid')
        title_c.alignment = Alignment(horizontal='center', vertical='center')
        ws_sum.row_dimensions[2].height = 38

        # Sous-titre
        ws_sum.merge_cells('B3:N3')
        st = ws_sum['B3']
        st.value = "Centre Julianna  |  Vue manager  |  30 derniers jours"
        st.font = Font(italic=True, size=10, color=C['text_lt'])
        st.alignment = Alignment(horizontal='center')
        ws_sum.row_dimensions[3].height = 16

        # KPI Row 1 — Finances
        _section_hdr(ws_sum, 5, "VALORISATION DU STOCK", 14, C['navy'])

        kpi1 = [
            ("Valeur d'achat stock",    f"{total_stock_value:,.0f} XAF",   C['blue']),
            ("Valeur de vente stock",   f"{total_retail_value:,.0f} XAF",  C['teal']),
            ("Marge potentielle",       f"{total_potential_margin:,.0f} XAF", C['green']),
            ("CA réalisé (30j)",        f"{total_revenue_30:,.0f} XAF",    C['purple']),
            ("Pertes 30j",              f"{total_loss_value:,.0f} XAF",    C['red']),
        ]
        cols_kpi = [2, 5, 8, 11, 14]
        for (lbl, val, bg), col in zip(kpi1, cols_kpi):
            ws_sum.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+2)
            ws_sum.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+2)
            _kpi_block(ws_sum, 6, col, lbl, val, bg)
        ws_sum.row_dimensions[6].height = 20
        ws_sum.row_dimensions[7].height = 26

        # KPI Row 2 — Stock
        _section_hdr(ws_sum, 9, "ÉTAT DU STOCK PHYSIQUE", 14, C['teal'])
        kpi2 = [
            ("Produits physiques",  str(len(physical)),   C['navy']),
            ("Stock OK",            str(ok_count),         C['green']),
            ("Stock bas",           str(low_count),        C['orange']),
            ("Ruptures",            str(out_count),        C['red']),
            ("Péremptions < 30j",   str(expiring_30 + batch_expiring), C['purple']),
        ]
        for (lbl, val, bg), col in zip(kpi2, cols_kpi):
            ws_sum.merge_cells(start_row=10, start_column=col, end_row=10, end_column=col+2)
            ws_sum.merge_cells(start_row=11, start_column=col, end_row=11, end_column=col+2)
            _kpi_block(ws_sum, 10, col, lbl, val, bg)
        ws_sum.row_dimensions[10].height = 20
        ws_sum.row_dimensions[11].height = 26

        # KPI Row 3 — Catalogue
        _section_hdr(ws_sum, 13, "CATALOGUE", 14, C['purple'])
        kpi3 = [
            ("Total produits",      str(len(all_products)),  C['navy']),
            ("Services",            str(len(services)),      C['blue']),
            ("Avec lots actifs",    str(sum(1 for p in physical if any(b.status in ('available','opened') for b in p.batches.all()))), C['teal']),
            ("Inactifs",            str(sum(1 for p in all_products if not p.is_active)), C['orange']),
            ("Produits vendus 30j", str(len(sales_map)),     C['green']),
        ]
        for (lbl, val, bg), col in zip(kpi3, cols_kpi):
            ws_sum.merge_cells(start_row=14, start_column=col, end_row=14, end_column=col+2)
            ws_sum.merge_cells(start_row=15, start_column=col, end_row=15, end_column=col+2)
            _kpi_block(ws_sum, 14, col, lbl, val, bg)
        ws_sum.row_dimensions[14].height = 20
        ws_sum.row_dimensions[15].height = 26

        # ─────────────────────────────────────────────────────────────────────
        # ONGLET 2 : PRODUITS PHYSIQUES
        # ─────────────────────────────────────────────────────────────────────
        ws_ph = wb.create_sheet("Produits Physiques")
        ws_ph.freeze_panes = 'A3'
        ws_ph.sheet_view.showGridLines = False

        PH_COLS = [
            ('Nom du produit',    38, 'left'),
            ('Référence',         16, 'center'),
            ('Catégorie',         20, 'center'),
            ('Fournisseur',       22, 'center'),
            ('Prix vente',        14, 'center'),
            ("Prix achat",        14, 'center'),
            ('Marge (XAF)',       14, 'center'),
            ('Marge %',           10, 'center'),
            ('Stock actuel',      13, 'center'),
            ('Valeur stock',      14, 'center'),
            ('Seuil bas',         10, 'center'),
            ('Statut',            12, 'center'),
            ('Ventes 30j (qté)',  14, 'center'),
            ('CA 30j (XAF)',      14, 'center'),
            ('Pertes 30j',        12, 'center'),
            ('Unité',             10, 'center'),
            ('Date péremption',   16, 'center'),
            ('Actif',              7, 'center'),
        ]

        _title_row(ws_ph, 1, f"PRODUITS PHYSIQUES  —  {today.strftime('%d/%m/%Y')}", len(PH_COLS))
        ws_ph.row_dimensions[1].height = 22

        for ci, (lbl, w, _) in enumerate(PH_COLS, 1):
            _hdr(ws_ph.cell(2, ci), lbl)
            ws_ph.column_dimensions[get_column_letter(ci)].width = w
        ws_ph.row_dimensions[2].height = 22

        # Données pour graphiques
        top_stock  = []   # (nom, stock)
        top_value  = []   # (nom, valeur_stock)
        top_margin = []   # (nom, marge%)
        top_sales  = []   # (nom, ca_30j)
        cat_value  = defaultdict(float)
        supplier_value = defaultdict(float)

        for ri, p in enumerate(physical, 3):
            pid   = str(p.id)
            stock = stock_cache[pid]
            thr   = p.low_stock_threshold or 0
            price = float(p.price or 0)
            cost  = float(p.cost_price or 0)
            margin_xaf = price - cost
            margin_pct = ((margin_xaf / cost) * 100) if cost > 0 else 0
            stock_val  = stock * cost
            s30 = sales_map.get(pid, {})
            l30 = loss_map.get(pid, {})
            qty_sold  = int(s30.get('qty') or 0)
            ca_30     = float(s30.get('revenue') or 0)
            loss_qty  = int(l30.get('qty') or 0)

            if stock <= 0:
                status, sc = 'RUPTURE',   C['red']
            elif stock <= thr:
                status, sc = 'STOCK BAS', C['orange']
            else:
                status, sc = 'OK',        C['green']

            name = p.name or ''
            cat  = p.category.name if p.category else 'Sans catégorie'
            sup  = p.supplier.name if p.supplier else ''
            row  = [
                name, p.reference or '', cat, sup,
                price, cost, margin_xaf, margin_pct / 100,
                stock, stock_val, thr, status,
                qty_sold, ca_30, loss_qty,
                p.get_base_unit_display() if hasattr(p, 'get_base_unit_display') else '',
                str(p.expiration_date) if p.expiration_date else '',
                'Oui' if p.is_active else 'Non',
            ]
            FMTS = [
                None, None, None, None,
                PRICE_FMT, PRICE_FMT, PRICE_FMT, '0.0%',
                INT_FMT, PRICE_FMT, INT_FMT, None,
                INT_FMT, PRICE_FMT, INT_FMT,
                None, None, None,
            ]
            for ci, (val, fmt, (_, _, align)) in enumerate(zip(row, FMTS, PH_COLS), 1):
                extra = {}
                if ci == 12:
                    extra = {'fg': sc, 'bold': True}
                elif ci == 7 and margin_xaf < 0:
                    extra = {'fg': C['red']}
                _cell(ws_ph.cell(ri, ci), val, ri, fmt=fmt, align=align, **extra)

            # Collecte graphiques
            if stock > 0:
                top_stock.append((name[:30], stock))
            top_value.append((name[:30], round(stock_val)))
            if margin_pct > 0:
                top_margin.append((name[:30], round(margin_pct, 1)))
            if ca_30 > 0:
                top_sales.append((name[:30], round(ca_30)))
            cat_value[cat] += stock_val
            supplier_value[sup or 'Inconnu'] += stock_val

        ws_ph.auto_filter.ref = f"A2:{get_column_letter(len(PH_COLS))}2"

        # ─────────────────────────────────────────────────────────────────────
        # ONGLET 3 : SERVICES
        # ─────────────────────────────────────────────────────────────────────
        ws_svc = wb.create_sheet("Services")
        ws_svc.freeze_panes = 'A3'
        ws_svc.sheet_view.showGridLines = False

        SVC_COLS = [
            ('Nom du service',   38, 'left'),
            ('Référence',        16, 'center'),
            ('Catégorie',        22, 'center'),
            ('Prix de vente',    16, 'center'),
            ("Prix d'achat",     16, 'center'),
            ('Marge (XAF)',      14, 'center'),
            ('Marge %',          10, 'center'),
            ('CA 30j (XAF)',     16, 'center'),
            ('Ventes 30j',       12, 'center'),
            ('Actif',             7, 'center'),
            ('Description',      40, 'left'),
        ]
        _title_row(ws_svc, 1, f"SERVICES  —  {today.strftime('%d/%m/%Y')}", len(SVC_COLS), C['teal'])
        ws_svc.row_dimensions[1].height = 22

        for ci, (lbl, w, _) in enumerate(SVC_COLS, 1):
            _hdr(ws_svc.cell(2, ci), lbl, bg=C['teal'])
            ws_svc.column_dimensions[get_column_letter(ci)].width = w
        ws_svc.row_dimensions[2].height = 22

        svc_top_sales = []
        for ri, p in enumerate(services, 3):
            pid   = str(p.id)
            price = float(p.price or 0)
            cost  = float(p.cost_price or 0)
            mg    = price - cost
            mgp   = ((mg / cost) * 100) if cost > 0 else 0
            s30   = sales_map.get(pid, {})
            ca_30 = float(s30.get('revenue') or 0)
            qty_s = int(s30.get('qty') or 0)

            row = [
                p.name or '', p.reference or '',
                p.category.name if p.category else '',
                price, cost, mg, mgp / 100,
                ca_30, qty_s,
                'Oui' if p.is_active else 'Non',
                p.description or '',
            ]
            FMTS_S = [None, None, None, PRICE_FMT, PRICE_FMT, PRICE_FMT, '0.0%',
                      PRICE_FMT, INT_FMT, None, None]
            for ci, (val, fmt, (_, _, align)) in enumerate(zip(row, FMTS_S, SVC_COLS), 1):
                extra = {}
                if ci == 6 and mg < 0:
                    extra = {'fg': C['red']}
                _cell(ws_svc.cell(ri, ci), val, ri, fmt=fmt, align=align, **extra)

            if ca_30 > 0:
                svc_top_sales.append(((p.name or '')[:30], round(ca_30)))

        ws_svc.auto_filter.ref = f"A2:{get_column_letter(len(SVC_COLS))}2"

        # ─────────────────────────────────────────────────────────────────────
        # ONGLET 4 : ALERTES
        # ─────────────────────────────────────────────────────────────────────
        ws_al = wb.create_sheet("Alertes")
        ws_al.sheet_view.showGridLines = False
        ws_al.column_dimensions['A'].width = 3

        _title_row(ws_al, 1, f"ALERTES CRITIQUES  —  {today.strftime('%d/%m/%Y')}", 10, C['red'])

        row_cur = 3

        # --- Ruptures ---
        ruptures = [p for p in physical if stock_cache[str(p.id)] <= 0]
        _section_hdr(ws_al, row_cur, f"RUPTURES DE STOCK ({len(ruptures)} produits)", 10, C['red'])
        row_cur += 1
        AL_HEADS = ['Produit', 'Référence', 'Catégorie', 'Fournisseur', 'Prix vente',
                    'Délai livraison (j)', 'Dernière vente 30j']
        ALWS = [35, 16, 20, 22, 14, 18, 18]
        for ci, (h, w) in enumerate(zip(AL_HEADS, ALWS), 2):
            _hdr(ws_al.cell(row_cur, ci), h, bg=C['red'])
            ws_al.column_dimensions[get_column_letter(ci)].width = w
        row_cur += 1
        for ri_off, p in enumerate(ruptures):
            pid = str(p.id)
            s30 = sales_map.get(pid, {})
            row = [p.name or '', p.reference or '',
                   p.category.name if p.category else '',
                   p.supplier.name if p.supplier else '',
                   float(p.price or 0),
                   p.supply_lead_time_days or '',
                   int(s30.get('qty') or 0)]
            for ci, v in enumerate(row, 2):
                fmt = PRICE_FMT if ci == 6 else (INT_FMT if ci in (7, 9) else None)
                _cell(ws_al.cell(row_cur, ci), v, row_cur, fmt=fmt,
                      bg=C['red_lt'])
            row_cur += 1

        row_cur += 1

        # --- Stock bas ---
        low_list = [p for p in physical
                    if 0 < stock_cache[str(p.id)] <= (p.low_stock_threshold or 0)]
        _section_hdr(ws_al, row_cur, f"STOCK BAS ({len(low_list)} produits)", 10, C['orange'])
        row_cur += 1
        SB_HEADS = ['Produit', 'Référence', 'Stock actuel', 'Seuil', 'Valeur restante',
                    'Ventes 30j', 'Jours restants estimés']
        SBW = [35, 16, 14, 10, 16, 12, 20]
        for ci, (h, w) in enumerate(zip(SB_HEADS, SBW), 2):
            _hdr(ws_al.cell(row_cur, ci), h, bg=C['orange'])
            ws_al.column_dimensions[get_column_letter(ci)].width = w
        row_cur += 1
        for p in low_list:
            pid   = str(p.id)
            stock = stock_cache[pid]
            s30   = sales_map.get(pid, {})
            qty30 = int(s30.get('qty') or 0)
            daily = qty30 / 30 if qty30 > 0 else 0
            days_left = round(stock / daily) if daily > 0 else '—'
            row = [p.name or '', p.reference or '',
                   stock, p.low_stock_threshold or 0,
                   stock * float(p.cost_price or 0),
                   qty30, days_left]
            for ci, v in enumerate(row, 2):
                fmt = PRICE_FMT if ci == 6 else (INT_FMT if ci in (4, 5, 7) else None)
                _cell(ws_al.cell(row_cur, ci), v, row_cur, fmt=fmt,
                      bg=C['orange_lt'])
            row_cur += 1

        row_cur += 1

        # --- Péremptions proches ---
        expiring = [p for p in physical
                    if p.expiration_date and p.expiration_date <= today + __import__('datetime').timedelta(days=60)]
        expiring.sort(key=lambda x: x.expiration_date)
        _section_hdr(ws_al, row_cur, f"PEREMPTIONS PROCHAINES — 60 jours ({len(expiring)} produits)", 10, C['purple'])
        row_cur += 1
        EX_HEADS = ['Produit', 'Référence', 'Date péremption', 'Jours restants',
                    'Stock', 'Valeur à risque']
        EXW = [35, 16, 18, 16, 12, 16]
        for ci, (h, w) in enumerate(zip(EX_HEADS, EXW), 2):
            _hdr(ws_al.cell(row_cur, ci), h, bg=C['purple'])
            ws_al.column_dimensions[get_column_letter(ci)].width = w
        row_cur += 1
        for p in expiring:
            stock = stock_cache[str(p.id)]
            days_r = (p.expiration_date - today).days
            val_risk = stock * float(p.cost_price or 0)
            bg_ex = C['red_lt'] if days_r <= 15 else C['orange_lt']
            row = [p.name or '', p.reference or '',
                   str(p.expiration_date), days_r, stock, val_risk]
            for ci, v in enumerate(row, 2):
                fmt = PRICE_FMT if ci == 7 else (INT_FMT if ci in (5, 6) else None)
                _cell(ws_al.cell(row_cur, ci), v, row_cur, fmt=fmt, bg=bg_ex)
            row_cur += 1

        # ─────────────────────────────────────────────────────────────────────
        # DONNÉES GRAPHIQUES (onglet caché)
        # ─────────────────────────────────────────────────────────────────────
        ws_d = wb.create_sheet("_Data")
        ws_d.sheet_state = 'hidden'

        def _write_block(ws, start_col, headers, rows):
            """Écrit un bloc (headers + data) à partir de start_col, retourne nb lignes"""
            for ci, h in enumerate(headers, start_col):
                ws.cell(1, ci, h)
            for ri, row in enumerate(rows, 2):
                for ci, v in enumerate(row, start_col):
                    ws.cell(ri, ci, v)
            return len(rows)

        # Col A-B : Top 15 stock
        top15s = sorted(top_stock, key=lambda x: x[1], reverse=True)[:15]
        n_ts = _write_block(ws_d, 1, ['Produit', 'Stock'], top15s)

        # Col D-E : Top 15 valeur stock
        top15v = sorted(top_value, key=lambda x: x[1], reverse=True)[:15]
        n_tv = _write_block(ws_d, 4, ['Produit', 'Valeur stock'], top15v)

        # Col G-H : Top 10 CA 30j (physiques)
        top10ca = sorted(top_sales, key=lambda x: x[1], reverse=True)[:10]
        n_ca = _write_block(ws_d, 7, ['Produit', 'CA 30j'], top10ca)

        # Col J-K : Top 10 marge %
        top10mg = sorted(top_margin, key=lambda x: x[1], reverse=True)[:10]
        n_mg = _write_block(ws_d, 10, ['Produit', 'Marge %'], top10mg)

        # Col M-N : Statut stock
        _write_block(ws_d, 13, ['Statut', 'Nb'],
                     [['OK', ok_count], ['Stock bas', low_count], ['Rupture', out_count]])

        # Col P-Q : Top 8 catégories par valeur
        top_cat_v = sorted(cat_value.items(), key=lambda x: x[1], reverse=True)[:8]
        n_catv = _write_block(ws_d, 16, ['Catégorie', 'Valeur stock'], top_cat_v)

        # Col S-T : Top 8 fournisseurs par valeur
        top_sup_v = sorted(supplier_value.items(), key=lambda x: x[1], reverse=True)[:8]
        n_supv = _write_block(ws_d, 19, ['Fournisseur', 'Valeur stock'], top_sup_v)

        # Col V-W : Top 10 CA services
        top_svc = sorted(svc_top_sales, key=lambda x: x[1], reverse=True)[:10]
        n_svc = _write_block(ws_d, 22, ['Service', 'CA 30j'], top_svc)

        # Col Y-Z : Jours de couverture stock (produits à risque, < 90 jours)
        couv_data = []
        for p in physical:
            pid  = str(p.id)
            stk  = stock_cache[pid]
            qty30 = int((sales_map.get(pid, {}).get('qty') or 0))
            if stk > 0 and qty30 > 0:
                jours = round(stk / (qty30 / 30))
                if jours <= 90:
                    couv_data.append(((p.name or '')[:32], jours))
        couv_data.sort(key=lambda x: x[1])          # du plus critique au moins critique
        top10_couv = couv_data[:10]
        n_couv = _write_block(ws_d, 25, ['Produit', 'Jours couverture'], top10_couv)

        # Col AB-AC : Top pertes 30j (valeur XAF)
        pertes_data = [
            ((p.name or '')[:32], round(float(loss_map[str(p.id)].get('val') or 0)))
            for p in physical
            if str(p.id) in loss_map
            and float(loss_map[str(p.id)].get('val') or 0) > 0
        ]
        pertes_data.sort(key=lambda x: x[1], reverse=True)
        top10_pertes = pertes_data[:10]
        n_pertes = _write_block(ws_d, 28, ['Produit', 'Pertes (XAF)'], top10_pertes)

        # ─────────────────────────────────────────────────────────────────────
        # ONGLET 5 : GRAPHIQUES — layout calculé précisément
        # ─────────────────────────────────────────────────────────────────────
        ws_g = wb.create_sheet("Graphiques")
        ws_g.sheet_view.showGridLines = False
        ws_g.sheet_view.showRowColHeaders = False

        # ── Calibration colonnes ─────────────────────────────────────────────
        # 1 unité de largeur Excel ≈ 0.185 cm (Calibri 11pt, 96 DPI)
        # On utilise COL_W=4 → ~0.74 cm/colonne
        COL_W     = 4
        CM_COL    = COL_W * 0.185        # ≈ 0.74 cm par colonne
        # 1 pt de hauteur de ligne ≈ 0.0353 cm → 15pt ≈ 0.529 cm/ligne
        CM_ROW    = 15 * 0.0353          # ≈ 0.529 cm par ligne (hauteur standard)

        # Toutes les colonnes uniformes (A inclus) pour positions reproductibles
        for c_i in range(1, 90):
            ws_g.column_dimensions[get_column_letter(c_i)].width = COL_W

        def _anchor(col_idx, row_idx):
            """Retourne la référence cellule ex. 'B5' pour col_idx=2, row_idx=5"""
            return f"{get_column_letter(col_idx)}{row_idx}"

        def _chart_col_span(width_cm):
            """Nombre de colonnes occupées par un graphique de width_cm cm"""
            return int(width_cm / CM_COL) + 1

        def _chart_row_span(height_cm):
            """Nombre de lignes occupées par un graphique de height_cm cm"""
            return int(height_cm / CM_ROW) + 1

        GAP_COLS = 3   # colonnes de séparation entre 2 graphiques côte à côte
        GAP_ROWS = 4   # lignes de séparation entre 2 rangées de graphiques

        def _section_banner(ws, row, col_end, text, bg, size=12):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=col_end)
            c = ws.cell(row, 1, text)
            c.font = Font(bold=True, size=size, color=C['white'])
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
            ws.row_dimensions[row].height = 26

        def _spacer(ws, row, height_pt=10):
            ws.row_dimensions[row].height = height_pt

        # ── Builders graphiques ───────────────────────────────────────────────
        from openpyxl.chart.label import DataLabelList

        def _labels(val_fmt='#,##0', show_pct=False, show_cat=False):
            """Crée un DataLabelList configuré"""
            dls = DataLabelList()
            dls.showVal      = not show_pct          # valeur brute sauf si camembert %
            dls.showPercent  = show_pct
            dls.showCatName  = show_cat
            dls.showSerName  = False
            dls.showLegendKey = False
            dls.numFmt       = val_fmt
            return dls

        def _style_series(series, color):
            """Applique couleur de remplissage + contour sur une série"""
            series.graphicalProperties.solidFill = color
            series.graphicalProperties.line.solidFill = color

        def _make_bar_h(title, data_col, cat_col, n, w=22, h=17,
                        color=None, val_fmt='#,##0'):
            """Barres horizontales avec étiquettes de valeurs"""
            ch = BarChart()
            ch.type    = 'bar'
            ch.style   = 10
            ch.title   = title
            ch.width   = w
            ch.height  = h
            ch.legend  = None
            # Grille légère uniquement sur l'axe des valeurs (x pour barres h)
            ch.y_axis.majorGridlines = None
            ch.x_axis.numFmt         = val_fmt
            ch.x_axis.majorGridlines = None
            ch.y_axis.tickLblPos     = 'low'    # labels catégories à gauche
            dr = Reference(ws_d, min_col=data_col, min_row=1, max_row=n + 1)
            cr = Reference(ws_d, min_col=cat_col,  min_row=2, max_row=n + 1)
            ch.add_data(dr, titles_from_data=True)
            ch.set_categories(cr)
            if ch.series:
                if color:
                    _style_series(ch.series[0], color)
                ch.series[0].dLbls = _labels(val_fmt)
            return ch

        def _make_bar_v(title, data_col, cat_col, n, w=22, h=17,
                        color=None, val_fmt='#,##0'):
            """Barres verticales avec étiquettes de valeurs"""
            ch = BarChart()
            ch.type    = 'col'
            ch.style   = 10
            ch.title   = title
            ch.width   = w
            ch.height  = h
            ch.legend  = None
            ch.x_axis.majorGridlines = None
            ch.y_axis.numFmt         = val_fmt
            ch.y_axis.majorGridlines = None
            dr = Reference(ws_d, min_col=data_col, min_row=1, max_row=n + 1)
            cr = Reference(ws_d, min_col=cat_col,  min_row=2, max_row=n + 1)
            ch.add_data(dr, titles_from_data=True)
            ch.set_categories(cr)
            if ch.series:
                if color:
                    _style_series(ch.series[0], color)
                ch.series[0].dLbls = _labels(val_fmt)
            return ch

        def _make_pie(title, data_col, cat_col, n, colors=None, w=18, h=17):
            """Camembert avec % + nom de catégorie sur chaque tranche"""
            ch = PieChart()
            ch.style  = 10
            ch.title  = title
            ch.width  = w
            ch.height = h
            dr = Reference(ws_d, min_col=data_col, min_row=1, max_row=n + 1)
            cr = Reference(ws_d, min_col=cat_col,  min_row=2, max_row=n + 1)
            ch.add_data(dr, titles_from_data=True)
            ch.set_categories(cr)
            if ch.series:
                if colors:
                    for idx, fill_color in enumerate(colors[:n]):
                        pt = DataPoint(idx=idx)
                        pt.graphicalProperties.solidFill = fill_color
                        ch.series[0].data_points.append(pt)
                # Étiquettes : nom + pourcentage sur chaque tranche
                dls = DataLabelList()
                dls.showPercent   = True
                dls.showCatName   = True
                dls.showVal       = False
                dls.showSerName   = False
                dls.showLegendKey = False
                dls.separator     = '\n'
                ch.series[0].dLbls = dls
            return ch

        # ── Dimensions des graphiques ─────────────────────────────────────────
        # Section 1 : 2 grands graphiques côte à côte  — 22cm × 15cm chacun
        W_BIG, H_BIG     = 22, 15
        # Section 2 : 2 graphiques côte à côte         — 22cm × 15cm
        # Section 3 : 1 camembert + 2 barres           — 16cm × 15cm / 22cm × 15cm
        W_PIE, H_PIE     = 16, 15
        W_MED, H_MED     = 22, 15
        # Section 4 : 1 graphique seul                 — 22cm × 15cm

        SPAN_BIG  = _chart_col_span(W_BIG)    # colonnes occupées par un grand graphique
        SPAN_PIE  = _chart_col_span(W_PIE)    # camembert
        SPAN_MED  = _chart_col_span(W_MED)    # moyen
        ROWS_BIG  = _chart_row_span(H_BIG)    # lignes occupées (≈ 29)

        START_COL = 1   # graphiques démarrent en colonne A

        # Positions colonne (0-based depuis START_COL)
        col_g1 = START_COL
        col_g2 = START_COL + SPAN_BIG + GAP_COLS

        col_g4 = START_COL
        col_g5 = START_COL + SPAN_BIG + GAP_COLS

        col_g3  = START_COL
        col_g6  = START_COL + SPAN_PIE + GAP_COLS
        col_g7  = START_COL + SPAN_PIE + GAP_COLS + SPAN_MED + GAP_COLS

        col_g8  = START_COL

        # Largeur totale du tableau de bord (pour les bandeaux)
        max_col = max(col_g2 + SPAN_BIG, col_g7 + SPAN_MED) + 2

        # ── Bandeau titre ────────────────────────────────────────────────────
        ws_g.merge_cells(start_row=1, start_column=1,
                         end_row=1, end_column=max_col)
        tc = ws_g.cell(1, 1,
            f"TABLEAU DE BORD INVENTAIRE  \u2014  {today.strftime('%d %B %Y').upper()}")
        tc.font      = Font(bold=True, size=22, color=C['white'])
        tc.fill      = PatternFill(start_color=C['navy'],
                                   end_color=C['navy'], fill_type='solid')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws_g.row_dimensions[1].height = 54

        ws_g.merge_cells(start_row=2, start_column=1,
                         end_row=2, end_column=max_col)
        sc = ws_g.cell(2, 1,
            f"Centre Julianna  |  Exporté le {today.strftime('%d/%m/%Y')}  "
            f"|  {len(physical)} produits physiques  |  {len(services)} services")
        sc.font      = Font(italic=True, size=10, color=C['text_lt'])
        sc.alignment = Alignment(horizontal='center', vertical='center')
        ws_g.row_dimensions[2].height = 20

        # ════════════════════════════════════════════════════════════════════
        # SECTION 1 — ANALYSE DU STOCK
        # ════════════════════════════════════════════════════════════════════
        row_s1 = 4
        _section_banner(ws_g, row_s1, max_col, "  SECTION 1  —  ANALYSE DU STOCK", C['navy'])
        _spacer(ws_g, row_s1 + 1, 12)
        row_charts_s1 = row_s1 + 2    # ligne de départ des graphiques section 1

        if n_ts > 0:
            ch_g1 = _make_bar_h(
                f"Top {n_ts} produits  —  Stock disponible (unités)",
                2, 1, n_ts, w=W_BIG, h=H_BIG, color=C['blue'])
            ws_g.add_chart(ch_g1, _anchor(col_g1, row_charts_s1))

        if n_tv > 0:
            ch_g2 = _make_bar_h(
                f"Top {n_tv} produits  —  Valeur du stock (XAF)",
                5, 4, n_tv, w=W_BIG, h=H_BIG, color=C['teal'])
            ws_g.add_chart(ch_g2, _anchor(col_g2, row_charts_s1))

        # ════════════════════════════════════════════════════════════════════
        # SECTION 2 — PERFORMANCE COMMERCIALE
        # ════════════════════════════════════════════════════════════════════
        row_s2 = row_charts_s1 + ROWS_BIG + GAP_ROWS
        _section_banner(ws_g, row_s2, max_col,
                        "  SECTION 2  —  PERFORMANCE COMMERCIALE (30 DERNIERS JOURS)",
                        C['purple'])
        _spacer(ws_g, row_s2 + 1, 12)
        row_charts_s2 = row_s2 + 2

        if n_ca > 0:
            ch_g4 = _make_bar_h(
                f"Top {n_ca} produits  —  Chiffre d'affaires 30j (XAF)",
                8, 7, n_ca, w=W_BIG, h=H_BIG, color=C['purple'])
            ws_g.add_chart(ch_g4, _anchor(col_g4, row_charts_s2))

        if n_mg > 0:
            ch_g5 = _make_bar_h(
                f"Top {n_mg} produits  —  Marge brute (%)",
                11, 10, n_mg, w=W_BIG, h=H_BIG, color=C['gold'])
            ws_g.add_chart(ch_g5, _anchor(col_g5, row_charts_s2))

        # ════════════════════════════════════════════════════════════════════
        # SECTION 3 — RÉPARTITION & ALERTES
        # ════════════════════════════════════════════════════════════════════
        row_s3 = row_charts_s2 + ROWS_BIG + GAP_ROWS
        _section_banner(ws_g, row_s3, max_col,
                        "  SECTION 3  —  RÉPARTITION PAR CATÉGORIE & ALERTES STOCK",
                        C['teal'])
        _spacer(ws_g, row_s3 + 1, 12)
        row_charts_s3 = row_s3 + 2

        ch_g3 = _make_pie(
            "Statut stock  —  OK / Bas / Rupture",
            14, 13, 3,
            colors=[C['green'], C['orange'], C['red']],
            w=W_PIE, h=H_PIE)
        ws_g.add_chart(ch_g3, _anchor(col_g3, row_charts_s3))

        if n_catv > 0:
            ch_g6 = _make_bar_v(
                f"Top {n_catv} catégories  —  Valeur du stock (XAF)",
                17, 16, n_catv, w=W_MED, h=H_MED, color=C['navy'])
            ws_g.add_chart(ch_g6, _anchor(col_g6, row_charts_s3))

        if n_supv > 0:
            ch_g7 = _make_bar_v(
                f"Top {n_supv} fournisseurs  —  Valeur du stock (XAF)",
                20, 19, n_supv, w=W_MED, h=H_MED, color=C['blue'])
            ws_g.add_chart(ch_g7, _anchor(col_g7, row_charts_s3))

        # ════════════════════════════════════════════════════════════════════
        # SECTION 4 — SERVICES
        # ════════════════════════════════════════════════════════════════════
        row_s4 = row_charts_s3 + ROWS_BIG + GAP_ROWS
        _section_banner(ws_g, row_s4, max_col,
                        "  SECTION 4  —  SERVICES  —  CHIFFRE D'AFFAIRES 30 JOURS",
                        C['green'])
        _spacer(ws_g, row_s4 + 1, 12)
        row_charts_s4 = row_s4 + 2

        if n_svc > 0:
            ch_g8 = _make_bar_h(
                f"Top {n_svc} services  —  CA 30 jours (XAF)",
                23, 22, n_svc, w=W_BIG, h=H_BIG, color=C['teal'])
            ws_g.add_chart(ch_g8, _anchor(col_g8, row_charts_s4))

        # ════════════════════════════════════════════════════════════════════
        # SECTION 5 — RISQUES : COUVERTURE STOCK & PERTES
        # ════════════════════════════════════════════════════════════════════
        row_s5 = row_charts_s4 + ROWS_BIG + GAP_ROWS
        _section_banner(ws_g, row_s5, max_col,
                        "  SECTION 5  —  GESTION DES RISQUES  —  "
                        "COUVERTURE STOCK & PERTES 30 JOURS",
                        C['red'])
        _spacer(ws_g, row_s5 + 1, 12)
        row_charts_s5 = row_s5 + 2

        col_g9  = START_COL
        col_g10 = START_COL + SPAN_BIG + GAP_COLS

        if n_couv > 0:
            ch_g9 = _make_bar_h(
                f"Jours de couverture stock  —  Top {n_couv} produits les + critiques",
                26, 25, n_couv, w=W_BIG, h=H_BIG,
                color=C['orange'], val_fmt='#,##0')
            ws_g.add_chart(ch_g9, _anchor(col_g9, row_charts_s5))

        if n_pertes > 0:
            ch_g10 = _make_bar_h(
                f"Top {n_pertes} produits  —  Valeur des pertes déclarées 30j (XAF)",
                29, 28, n_pertes, w=W_BIG, h=H_BIG, color=C['red'])
            ws_g.add_chart(ch_g10, _anchor(col_g10, row_charts_s5))

        # Ligne de fin (espace blanc après le dernier graphique)
        _spacer(ws_g, row_charts_s5 + ROWS_BIG + 2, 20)

        # ── Mise en ordre ────────────────────────────────────────────────────
        wb.active = wb['Résumé Exécutif']

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="inventaire.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def expired_report(self, request):
        """Rapport des produits et lots périmés ou expirant bientôt"""
        from apps.invoicing.models import ProductBatch
        from django.utils import timezone
        today = timezone.now().date()
        
        # 1. Produits avec date de péremption globale dépassée
        expired_products = self.get_queryset().filter(
            product_type='physical',
            expiration_date__lte=today,
            stock_quantity__gt=0
        )
        
        # 2. Lots spécifiques périmés
        expired_batches = ProductBatch.objects.filter(
            organization=request.user.organization,
            expiry_date__lte=today,
            quantity_remaining__gt=0
        ).select_related('product')
        
        results = {
            'summary': {
                'total_expired_products': expired_products.count(),
                'total_expired_batches': expired_batches.count(),
                'estimated_loss_value': 0,
            },
            'expired_products': [],
            'expired_batches': []
        }
        
        total_loss = 0
        
        for p in expired_products:
            loss = float(p.stock_quantity * (p.cost_price or 0))
            total_loss += loss
            results['expired_products'].append({
                'id': str(p.id),
                'name': p.name,
                'reference': p.reference,
                'stock': p.stock_quantity,
                'expiry_date': p.expiration_date,
                'loss_value': loss
            })
            
        for b in expired_batches:
            loss = float(b.quantity_remaining * (b.product.cost_price or 0))
            total_loss += loss
            results['expired_batches'].append({
                'id': str(b.id),
                'product_name': b.product.name,
                'batch_number': b.batch_number,
                'quantity': b.quantity_remaining,
                'expiry_date': b.expiry_date,
                'loss_value': loss
            })
            
        results['summary']['estimated_loss_value'] = total_loss
        return Response(results)

    @action(detail=False, methods=['get'])
    def batch_stats(self, request):
        """Stats sur les lots pour les cartes de statistiques de la page produits"""
        from apps.invoicing.models import ProductBatch
        from django.utils import timezone
        from datetime import timedelta
        today = timezone.now().date()
        thirty_days = today + timedelta(days=30)
        sixty_days = today + timedelta(days=60)

        org = request.user.organization
        active_statuses = ['available', 'opened']

        expired_batches = ProductBatch.objects.filter(
            organization=org,
            expiry_date__lt=today,
            quantity_remaining__gt=0,
            status__in=active_statuses,
        ).count()

        expiring_30 = ProductBatch.objects.filter(
            organization=org,
            expiry_date__gte=today,
            expiry_date__lte=thirty_days,
            quantity_remaining__gt=0,
            status__in=active_statuses,
        ).count()

        expiring_60 = ProductBatch.objects.filter(
            organization=org,
            expiry_date__gt=thirty_days,
            expiry_date__lte=sixty_days,
            quantity_remaining__gt=0,
            status__in=active_statuses,
        ).count()

        total_batches = ProductBatch.objects.filter(
            organization=org,
            status__in=active_statuses,
            quantity_remaining__gt=0,
        ).count()

        return Response({
            'expired_batches': expired_batches,
            'expiring_soon_30': expiring_30,
            'expiring_soon_60': expiring_60,
            'total_active_batches': total_batches,
        })

    @action(detail=True, methods=['get'])
    def stock_movements(self, request, pk=None):
        """Historique des mouvements de stock pour un produit"""
        from apps.invoicing.models import StockMovement
        from .serializers import StockMovementSerializer

        product = self.get_object()
        movements = StockMovement.objects.filter(product=product).order_by('-created_at')

        # Pagination
        page = self.paginate_queryset(movements)
        if page is not None:
            serializer = StockMovementSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = StockMovementSerializer(movements, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='pdf-report')
    def generate_pdf_report(self, request, pk=None):
        """Générer un rapport PDF pour un produit"""
        from django.http import HttpResponse
        import traceback
        import re

        try:
            product = self.get_object()

            # Générer le PDF avec WeasyPrint
            from .services.report_generator_weasy import generate_product_report_pdf
            pdf_buffer = generate_product_report_pdf(product, request.user)

            # Lire le contenu du buffer
            pdf_content = pdf_buffer.getvalue()

            # Créer la réponse HTTP
            response = HttpResponse(pdf_content, content_type='application/pdf')

            product_name = getattr(product, 'name', 'produit') or 'produit'
            safe_filename = re.sub(r'[^\w\s-]', '', product_name).strip()
            safe_filename = re.sub(r'[-\s]+', '_', safe_filename)
            filename = f"rapport-produit-{safe_filename}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(pdf_content)

            return response

        except ImportError as e:
            return Response(
                {'error': 'Service de génération PDF non disponible', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='bulk-pdf-report')
    def generate_bulk_pdf_report(self, request):
        """Générer un rapport PDF pour plusieurs produits (avec filtres)"""
        from django.http import HttpResponse
        from datetime import datetime
        import traceback
        
        try:
            # Récupérer les paramètres de filtrage
            product_ids = request.data.get('product_ids', [])
            date_start = request.data.get('date_start')
            date_end = request.data.get('date_end')
            category_filter = request.data.get('category')
            
            # Construire le queryset
            queryset = self.get_queryset()
            
            # Filtrer par IDs si fournis
            if product_ids and len(product_ids) > 0:
                queryset = queryset.filter(id__in=product_ids)
            
            # Filtrer par dates (sur created_at)
            date_start_obj = None
            date_end_obj = None
            if date_start:
                try:
                    if isinstance(date_start, str):
                        date_start = date_start.replace('Z', '+00:00') if 'Z' in date_start else date_start
                        date_start_obj = datetime.fromisoformat(date_start.replace('Z', ''))
                    else:
                        date_start_obj = date_start
                    queryset = queryset.filter(created_at__gte=date_start_obj)
                except Exception as e:
                    print(f"Erreur parsing date_start: {e}")
            
            if date_end:
                try:
                    if isinstance(date_end, str):
                        date_end = date_end.replace('Z', '+00:00') if 'Z' in date_end else date_end
                        date_end_obj = datetime.fromisoformat(date_end.replace('Z', ''))
                    else:
                        date_end_obj = date_end
                    queryset = queryset.filter(created_at__lte=date_end_obj)
                except Exception as e:
                    print(f"Erreur parsing date_end: {e}")
            
            # Filtrer par catégorie
            if category_filter:
                queryset = queryset.filter(category_id=category_filter)
            
            # Vérifier qu'il y a des produits
            count = queryset.count()
            if count == 0:
                return Response(
                    {'error': 'Aucun produit trouvé avec les filtres spécifiés'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Limiter le nombre de produits (sécurité)
            queryset = queryset[:500]
            
            # Générer le PDF
            from .services.report_generator_weasy import generate_products_report_pdf
            pdf_buffer = generate_products_report_pdf(
                queryset,
                request.user,
                date_start_obj,
                date_end_obj
            )
            
            # Créer la réponse HTTP
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            
            filename = f"rapport-produits-{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(response.content)
            
            return response
            
        except ImportError as e:
            print(f"ImportError génération PDF produits: {e}")
            traceback.print_exc()
            return Response(
                {'error': 'Service de génération PDF non disponible', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            print(f"Erreur génération PDF produits: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du rapport: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def adjust_stock(self, request, pk=None):
        """Ajustement manuel du stock — supporte batch_id pour traçabilité par lot"""
        product = self.get_object()

        if product.product_type != 'physical':
            return Response(
                {'error': 'Seuls les produits physiques ont un stock'},
                status=status.HTTP_400_BAD_REQUEST
            )

        quantity = request.data.get('quantity')
        notes = request.data.get('notes', '')
        batch_id = request.data.get('batch_id')
        movement_type_param = request.data.get('movement_type', 'adjustment')

        valid_types = ['adjustment', 'reception', 'return', 'loss', 'initial']
        if movement_type_param not in valid_types:
            movement_type_param = 'adjustment'

        if quantity is None:
            return Response(
                {'error': 'La quantité est requise'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            quantity = int(quantity)
        except ValueError:
            return Response(
                {'error': 'Quantité invalide'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Résoudre le lot si batch_id fourni
        from apps.invoicing.models import StockMovement, ProductBatch
        batch = None
        if batch_id:
            try:
                batch = ProductBatch.objects.get(
                    id=batch_id, product=product,
                    organization=request.user.organization
                )
            except ProductBatch.DoesNotExist:
                return Response({'error': 'Lot introuvable'}, status=status.HTTP_404_NOT_FOUND)

            # Mettre à jour la quantité du lot
            if quantity < 0:
                abs_qty = abs(quantity)
                if batch.quantity_remaining < abs_qty:
                    return Response(
                        {'error': f'Stock insuffisant dans ce lot ({batch.quantity_remaining} disponible(s))'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                batch.quantity_remaining -= abs_qty
            else:
                batch.quantity_remaining += quantity
            
            batch.update_status()

        # Ajuster le stock produit
        old_quantity = product.stock_quantity
        product.stock_quantity += quantity
        product.save(update_fields=['stock_quantity'])

        movement = StockMovement.objects.create(
            product=product,
            batch=batch,
            movement_type=movement_type_param,
            quantity=quantity,
            quantity_before=old_quantity,
            quantity_after=product.stock_quantity,
            reference_type='manual',
            notes=notes,
            created_by=request.user if request.user.is_authenticated else None
        )

        from apps.invoicing.stock_alerts import check_stock_after_movement
        check_stock_after_movement(product)

        from .serializers import StockMovementSerializer
        return Response({
            'product': self.get_serializer(product).data,
            'movement': StockMovementSerializer(movement).data
        })

    @action(detail=True, methods=['post'])
    def report_loss(self, request, pk=None):
        """Déclarer une perte de stock avec raison détaillée"""
        product = self.get_object()

        if product.product_type != 'physical':
            return Response(
                {'error': 'Seuls les produits physiques peuvent avoir des pertes'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from apps.invoicing.models import StockMovement

        # Validation des données
        quantity = request.data.get('quantity')
        loss_reason = request.data.get('loss_reason')
        loss_description = request.data.get('loss_description', '')
        notes = request.data.get('notes', '')

        if not quantity:
            return Response({'error': 'La quantité est requise'}, status=status.HTTP_400_BAD_REQUEST)

        if not loss_reason:
            return Response({'error': 'La raison de la perte est requise'}, status=status.HTTP_400_BAD_REQUEST)

        if loss_reason not in dict(StockMovement.LOSS_REASONS):
            return Response({'error': 'Raison de perte invalide'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            quantity_lost = abs(int(quantity))
        except ValueError:
            return Response({'error': 'Quantité invalide'}, status=status.HTTP_400_BAD_REQUEST)

        # Calculer la valeur de la perte
        loss_value = product.cost_price * quantity_lost if product.cost_price else 0

        # Créer le mouvement de perte
        old_quantity = product.stock_quantity
        product.stock_quantity -= quantity_lost
        product.save(update_fields=['stock_quantity'])

        movement = StockMovement.objects.create(
            product=product,
            movement_type='loss',
            quantity=-quantity_lost,
            quantity_before=old_quantity,
            quantity_after=product.stock_quantity,
            reference_type='loss_report',
            loss_reason=loss_reason,
            loss_description=loss_description,
            loss_value=loss_value,
            notes=notes,
            created_by=request.user if request.user.is_authenticated else None
        )

        # Vérifier et envoyer alertes si nécessaire
        from apps.invoicing.stock_alerts import check_stock_after_movement
        check_stock_after_movement(product)

        from .serializers import StockMovementSerializer
        return Response({
            'product': self.get_serializer(product).data,
            'movement': StockMovementSerializer(movement).data,
            'loss_value': float(loss_value)
        })

    @action(detail=False, methods=['get'])
    def stock_alerts(self, request):
        """Liste des produits nécessitant une attention (stock bas/rupture)"""
        from apps.invoicing.stock_alerts import StockAlertService

        low_stock = StockAlertService.check_low_stock_products()
        out_of_stock = StockAlertService.get_out_of_stock_products()

        return Response({
            'low_stock': self.get_serializer(low_stock, many=True).data,
            'out_of_stock': self.get_serializer(out_of_stock, many=True).data,
            'low_stock_count': len(low_stock),
            'out_of_stock_count': len(out_of_stock)
        })
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Statistiques complètes inter-modules pour un produit"""
        product = self.get_object()
        from django.db.models import Sum, Count, Avg, Min, Max
        from datetime import timedelta
        from django.utils import timezone
        
        # Stats ventes (InvoiceItems)
        invoice_items = product.invoice_items.all()
        invoices = invoice_items.values('invoice').distinct()
        
        total_sales = invoice_items.aggregate(
            total_quantity=Sum('quantity'),
            total_amount=Sum('total_price')
        )
        
        # Top clients - Reformater pour le frontend
        top_clients_qs = invoice_items.filter(
            invoice__client__isnull=False
        ).values(
            'invoice__client__id',
            'invoice__client__name',
            'invoice__client__email'
        ).annotate(
            total_purchases=Sum('total_price'),
            purchase_count=Count('invoice', distinct=True),
            total_quantity=Sum('quantity')
        ).order_by('-total_purchases')[:10]
        
        # Convertir au format attendu par le frontend
        top_clients = [{
            'invoice__client__id': str(item['invoice__client__id']),
            'invoice__client__name': item['invoice__client__name'],
            'invoice__client__email': item['invoice__client__email'] or '',
            'total_purchases': float(item['total_purchases']) if item['total_purchases'] else 0,
            'purchase_count': item['purchase_count'],
            'total_quantity': item['total_quantity']
        } for item in top_clients_qs]
        
        # Factures récentes avec clients
        recent_invoices = invoice_items.select_related('invoice', 'invoice__client').order_by('-created_at')[:10]
        recent_invoices_data = []
        for item in recent_invoices:
            # Récupérer le nom du client avec fallback
            client_name = 'Aucun client'
            if item.invoice.client:
                client_name = item.invoice.client.name or 'Client sans nom'
            
            recent_invoices_data.append({
                'invoice_id': str(item.invoice.id),
                'invoice_number': item.invoice.invoice_number,
                'client_name': client_name,
                'created_at': item.invoice.created_at,
                'quantity': item.quantity,
                'total_price': float(item.total_price)
            })
        
        # Stats achats (PurchaseOrderItems)
        po_items = product.purchase_order_items.all()
        purchase_stats = po_items.aggregate(
            total_pos=Count('purchase_order', distinct=True),
            total_quantity=Sum('quantity'),
            avg_price=Avg('unit_price')
        )
        
        # Stats contrats
        contract_items = product.contract_items.filter(contract__status='active')
        contract_stats = contract_items.aggregate(
            active_contracts=Count('contract', distinct=True),
            min_price=Min('contracted_price'),
            max_price=Max('contracted_price')
        )
        
        # Stats e-sourcing
        bid_items = product.bid_items.all()
        sourcing_stats = {
            'total_bids': bid_items.count(),
            'awarded_bids': bid_items.filter(bid__status='awarded').count()
        }
        
        # Tendance (30 derniers jours vs 30 jours précédents)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        sixty_days_ago = timezone.now() - timedelta(days=60)
        
        recent_sales = invoice_items.filter(created_at__gte=thirty_days_ago).aggregate(
            total=Sum('total_price')
        )['total'] or 0
        
        previous_sales = invoice_items.filter(
            created_at__gte=sixty_days_ago,
            created_at__lt=thirty_days_ago
        ).aggregate(total=Sum('total_price'))['total'] or 0
        
        trend = ((recent_sales - previous_sales) / previous_sales * 100) if previous_sales > 0 else 0
        
        return Response({
            'product_id': str(product.id),
            'product_name': product.name,
            'product_reference': product.reference,
            
            'sales_summary': {
                'total_invoices': invoices.count(),
                'total_quantity_sold': int(total_sales['total_quantity'] or 0),
                'total_sales_amount': float(total_sales['total_amount'] or 0),
                'average_sale': float(total_sales['total_amount'] / invoices.count()) if invoices.count() > 0 else 0,
                'unique_clients': invoice_items.values('invoice__client').distinct().count(),
            },
            
            'purchase_summary': {
                'total_purchase_orders': purchase_stats['total_pos'] or 0,
                'total_quantity_purchased': int(purchase_stats['total_quantity'] or 0),
                'average_purchase_price': float(purchase_stats['avg_price'] or 0),
            },
            
            'contract_summary': {
                'active_contracts': contract_stats['active_contracts'] or 0,
                'contracted_price_min': float(contract_stats['min_price'] or 0),
                'contracted_price_max': float(contract_stats['max_price'] or 0),
            },
            
            'sourcing_summary': sourcing_stats,
            
            'warehouse_info': {
                'warehouse_id': str(product.warehouse.id) if product.warehouse else None,
                'warehouse_name': product.warehouse.name if product.warehouse else None,
                'warehouse_code': product.warehouse.code if product.warehouse else None,
                'location': f"{product.warehouse.city}, {product.warehouse.province}" if product.warehouse else None,
                'current_stock': product.stock_quantity,
                'stock_status': product.stock_status,
            },
            
            'top_clients': list(top_clients),
            'recent_invoices': recent_invoices_data,
            
            'sales_trend': {
                'last_30_days': float(recent_sales),
                'previous_30_days': float(previous_sales),
                'trend_percent': round(trend, 2),
            },
            
            'stock_movements_count': product.stock_movements.count(),
        })


class HasClientOrPatientModuleAccess(permissions.BasePermission):
    """
    Custom permission to allow access if user has either CLIENTS or PATIENTS module.
    """
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        
        has_clients = user_has_module_access(request.user, Modules.CLIENTS)
        has_patients = user_has_module_access(request.user, Modules.PATIENTS)
        
        return has_clients or has_patients

    def has_object_permission(self, request, view, obj):
        return self.has_permission(request, view)


class ClientViewSet(OrganizationFilterMixin, viewsets.ModelViewSet):
    """ViewSet pour les clients"""
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated, HasClientOrPatientModuleAccess]
    # required_module = Modules.CLIENTS  # Handled by custom permission
    organization_field = 'organization'  # Client has organization FK
    filterset_fields = ['is_active', 'client_type']
    search_fields = ['name', 'email', 'contact_person', 'phone']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Statistiques complètes pour un client"""
        client = self.get_object()
        from django.db.models import Sum, Count, Avg
        from datetime import timedelta
        from django.utils import timezone
        
        # Stats factures
        invoices = client.invoices.all()
        
        invoice_stats = {
            'draft': invoices.filter(status='draft').count(),
            'sent': invoices.filter(status='sent').count(),
            'paid': invoices.filter(status='paid').count(),
            'overdue': invoices.filter(status='overdue').count(),
            'cancelled': invoices.filter(status='cancelled').count(),
        }
        
        # Montants
        total_sales = invoices.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_paid = invoices.filter(status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_outstanding = invoices.filter(status__in=['sent', 'overdue']).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # Produits les plus achetés (avec gestion FK product ou champs texte)
        from apps.invoicing.models import InvoiceItem
        
        # Récupérer tous les items du client
        all_items = InvoiceItem.objects.filter(
            invoice__client=client
        ).select_related('product').order_by('-total_price')
        
        # Construire manuellement la liste avec fallbacks
        products_dict = {}
        
        for item in all_items:
            # Identifier le produit par son ID FK s'il existe, sinon par product_reference
            if item.product_id:
                product_key = str(item.product_id)
                product_name = item.product.name
                product_ref = item.product.reference
            else:
                # Pas de FK, utiliser les champs texte
                product_key = f"manual_{item.product_reference}_{item.description[:20]}"
                product_name = item.description or 'Produit sans nom'
                product_ref = item.product_reference or 'N/A'
            
            # Agréger les quantités pour le même produit
            if product_key in products_dict:
                products_dict[product_key]['total_quantity'] += item.quantity
                products_dict[product_key]['total_amount'] += float(item.total_price)
                products_dict[product_key]['purchase_count'] += 1
            else:
                products_dict[product_key] = {
                    'product__id': str(item.product_id) if item.product_id else None,
                    'product__name': product_name,
                    'product__reference': product_ref,
                    'total_quantity': item.quantity,
                    'total_amount': float(item.total_price),
                    'purchase_count': 1
                }
        
        # Convertir en liste et trier par montant total
        top_products = sorted(
            products_dict.values(),
            key=lambda x: x['total_amount'],
            reverse=True
        )[:10]
        
        # Factures récentes
        recent_invoices = invoices.order_by('-created_at')[:10]
        recent_invoices_data = [{
            'id': str(inv.id),
            'invoice_number': inv.invoice_number,
            'title': inv.title,
            'status': inv.status,
            'total_amount': float(inv.total_amount),
            'created_at': inv.created_at,
            'due_date': inv.due_date,
        } for inv in recent_invoices]
        
        # Tendance (30 derniers jours)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        sixty_days_ago = timezone.now() - timedelta(days=60)
        
        recent_sales = invoices.filter(created_at__gte=thirty_days_ago).aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        previous_sales = invoices.filter(
            created_at__gte=sixty_days_ago,
            created_at__lt=thirty_days_ago
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        trend = ((recent_sales - previous_sales) / previous_sales * 100) if previous_sales > 0 else 0
        
        return Response({
            'client_id': str(client.id),
            'client_name': client.name,
            
            'invoice_summary': {
                'total_invoices': invoices.count(),
                'total_sales_amount': float(total_sales),
                'total_paid_amount': float(total_paid),
                'total_outstanding': float(total_outstanding),
                'average_invoice_amount': float(total_sales / invoices.count()) if invoices.count() > 0 else 0,
            },
            
            'invoice_status_breakdown': invoice_stats,
            
            'top_products': list(top_products),
            'recent_invoices': recent_invoices_data,
            
            'sales_trend': {
                'last_30_days': float(recent_sales),
                'previous_30_days': float(previous_sales),
                'trend_percent': round(trend, 2),
            },
            
            'payment_info': {
                'payment_terms': client.payment_terms,
                'tax_id': client.tax_id,
            },
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export des clients en CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="clients.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Name', 'Contact Person', 'Email', 'Phone', 'Address', 'Payment Terms', 'Active'])
        
        for client in self.get_queryset():
            writer.writerow([
                client.name or '',
                client.contact_person or '',
                client.email or '',
                client.phone or '',
                client.address or '',
                client.payment_terms or '',
                'Oui' if client.is_active else 'Non'
            ])
        
        return response
    
    @action(detail=True, methods=['get'], url_path='pdf-report')
    def generate_pdf_report(self, request, pk=None):
        """Générer un rapport PDF pour un client"""
        from django.http import HttpResponse
        import traceback
        
        try:
            client = self.get_object()
            
            # Générer le PDF avec WeasyPrint
            from .services.report_generator_weasy import generate_client_report_pdf
            pdf_buffer = generate_client_report_pdf(client, request.user)
            
            # Créer la réponse HTTP
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            
            client_name = getattr(client, 'name', 'client') or 'client'
            filename = f"rapport-client-{client_name.replace(' ', '_')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(response.content)
            
            return response
            
        except ImportError as e:
            print(f"ImportError génération PDF client: {e}")
            traceback.print_exc()
            return Response(
                {'error': 'Service de génération PDF non disponible', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            print(f"Erreur génération PDF client: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du PDF: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='bulk-pdf-report')
    def generate_bulk_pdf_report(self, request):
        """Générer un rapport PDF pour plusieurs clients (avec filtres)"""
        from django.http import HttpResponse
        from datetime import datetime
        import traceback
        
        try:
            # Récupérer les paramètres de filtrage
            client_ids = request.data.get('client_ids', [])
            date_start = request.data.get('date_start')
            date_end = request.data.get('date_end')
            status_filter = request.data.get('status')
            
            # Construire le queryset
            queryset = self.get_queryset()
            
            # Filtrer par IDs si fournis
            if client_ids and len(client_ids) > 0:
                queryset = queryset.filter(id__in=client_ids)
            
            # Filtrer par dates (sur created_at)
            date_start_obj = None
            date_end_obj = None
            if date_start:
                try:
                    if isinstance(date_start, str):
                        date_start = date_start.replace('Z', '+00:00') if 'Z' in date_start else date_start
                        date_start_obj = datetime.fromisoformat(date_start.replace('Z', ''))
                    else:
                        date_start_obj = date_start
                    queryset = queryset.filter(created_at__gte=date_start_obj)
                except Exception as e:
                    print(f"Erreur parsing date_start: {e}")
            
            if date_end:
                try:
                    if isinstance(date_end, str):
                        date_end = date_end.replace('Z', '+00:00') if 'Z' in date_end else date_end
                        date_end_obj = datetime.fromisoformat(date_end.replace('Z', ''))
                    else:
                        date_end_obj = date_end
                    queryset = queryset.filter(created_at__lte=date_end_obj)
                except Exception as e:
                    print(f"Erreur parsing date_end: {e}")
            
            # Filtrer par statut
            if status_filter:
                queryset = queryset.filter(is_active=(status_filter == 'active'))
            
            # Vérifier qu'il y a des clients
            count = queryset.count()
            if count == 0:
                return Response(
                    {'error': 'Aucun client trouvé avec les filtres spécifiés'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Limiter le nombre de clients (sécurité)
            queryset = queryset[:500]
            
            # Générer le PDF
            from .services.report_generator_weasy import generate_clients_report_pdf
            pdf_buffer = generate_clients_report_pdf(
                queryset,
                request.user,
                date_start_obj,
                date_end_obj
            )
            
            # Créer la réponse HTTP
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            
            filename = f"rapport-clients-{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(response.content)
            
            return response
            
        except ImportError as e:
            print(f"ImportError génération PDF clients: {e}")
            traceback.print_exc()
            return Response(
                {'error': 'Service de génération PDF non disponible', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            print(f"Erreur génération PDF clients: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du rapport: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PurchaseOrderViewSet(OrganizationFilterMixin, viewsets.ModelViewSet):
    """ViewSet pour les bons de commande"""
    queryset = PurchaseOrder.objects.all()
    serializer_class = PurchaseOrderSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    required_module = Modules.PURCHASE_ORDERS
    organization_field = 'created_by__organization'  # Filter via user's org
    filterset_fields = ['status', 'supplier', 'created_by']
    search_fields = ['po_number', 'title', 'description']
    ordering_fields = ['created_at', 'total_amount', 'required_date']
    ordering = ['-created_at']
    
    def get_queryset(self):
        # First apply organization filter
        queryset = super().get_queryset()
        
        # Filtre par dates
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        """Ajouter un item au bon de commande"""
        purchase_order = self.get_object()
        serializer = PurchaseOrderItemSerializer(data=request.data)
        
        if serializer.is_valid():
            serializer.save(purchase_order=purchase_order)
            purchase_order.recalculate_totals()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approuver un bon de commande"""
        purchase_order = self.get_object()
        if purchase_order.status == 'draft':
            purchase_order.status = 'approved'
            purchase_order.approved_by = request.user if request.user.is_authenticated else None
            purchase_order.save()
            serializer = self.get_serializer(purchase_order)
            return Response(serializer.data)
        return Response(
            {'error': 'Only draft orders can be approved'},
            status=status.HTTP_400_BAD_REQUEST
        )

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Marquer le bon de commande comme reçu et mettre à jour le stock"""
        purchase_order = self.get_object()

        if purchase_order.status == 'received':
            return Response(
                {'error': 'Ce bon de commande a déjà été reçu'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Recevoir les items et mettre à jour le stock
        result = purchase_order.receive_items(
            user=request.user if request.user.is_authenticated else None
        )

        if 'error' in result:
            return Response({'error': result['error']}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(purchase_order)
        return Response({
            'purchase_order': serializer.data,
            'stock_update': result
        })

    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Envoyer un bon de commande par email"""
        purchase_order = self.get_object()

        # Email du destinataire (par défaut: fournisseur, sinon fourni dans la requête)
        recipient_email = request.data.get('recipient_email') or request.data.get('email')
        template_type = request.data.get('template', 'modern')
        custom_message = request.data.get('custom_message')
        
        # Récupérer la langue depuis la requête (header Accept-Language ou paramètre)
        language = request.data.get('language')
        if not language:
            # Essayer de récupérer depuis les headers HTTP
            accept_language = request.META.get('HTTP_ACCEPT_LANGUAGE', '')
            if 'en' in accept_language.lower():
                language = 'en'
            else:
                language = getattr(request, 'LANGUAGE_CODE', 'fr')

        # Envoyer l'email avec le PDF
        from .services.email_service import PurchaseOrderEmailService

        result = PurchaseOrderEmailService.send_purchase_order_email(
            po=purchase_order,
            recipient_email=recipient_email,
            template_type=template_type,
            custom_message=custom_message,
            language=language
        )

        if result['success']:
            serializer = self.get_serializer(purchase_order)
            return Response({
                'purchase_order': serializer.data,
                'message': result['message']
            })
        else:
            return Response(
                {'error': result['message']},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'], url_path='pdf')
    def generate_pdf(self, request, pk=None):
        """Générer PDF du bon de commande avec WeasyPrint"""
        from django.http import HttpResponse
        from io import BytesIO

        try:
            po = self.get_object()
            template_type = request.query_params.get('template', 'modern')

            # Utiliser la vue WeasyPrint existante
            try:
                from django.test import RequestFactory
                from apps.purchase_orders.views_pdf import PurchaseOrderPDFView

                # Créer une requête factice pour la vue
                factory = RequestFactory()
                fake_request = factory.get(f'/purchase-orders/{po.id}/pdf/?template={template_type}')
                fake_request.user = request.user

                # Appeler la vue pour générer le PDF
                view = PurchaseOrderPDFView()
                view.request = fake_request
                view.kwargs = {'pk': str(po.id)}

                response = view.get(fake_request, pk=str(po.id))

                if response.status_code == 200:
                    # Rendre la réponse WeasyPrint avant d'accéder au contenu
                    response.render()

                    # Retourner la réponse PDF
                    http_response = HttpResponse(
                        response.content,
                        content_type='application/pdf'
                    )
                    filename = f"bon-commande-{po.po_number}.pdf"
                    http_response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    http_response['Content-Length'] = len(response.content)
                    return http_response
                else:
                    raise Exception(f"Error generating PDF: HTTP {response.status_code}")

            except ImportError as import_error:
                return Response(
                    {'error': 'WeasyPrint not available', 'details': str(import_error)},
                    status=status.HTTP_503_SERVICE_UNAVAILABLE
                )

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='bulk-pdf-report')
    def generate_bulk_pdf_report(self, request):
        """Générer un rapport PDF pour plusieurs bons de commande (avec filtres)"""
        from django.http import HttpResponse
        from datetime import datetime
        import traceback
        
        try:
            # Récupérer les paramètres de filtrage
            po_ids = request.data.get('po_ids', [])
            date_start = request.data.get('date_start')
            date_end = request.data.get('date_end')
            status_filter = request.data.get('status')
            supplier_id = request.data.get('supplier_id')
            
            # Construire le queryset
            queryset = self.get_queryset()
            
            # Filtrer par IDs si fournis
            if po_ids and len(po_ids) > 0:
                queryset = queryset.filter(id__in=po_ids)
            
            # Filtrer par dates
            date_start_obj = None
            date_end_obj = None
            if date_start:
                try:
                    if isinstance(date_start, str):
                        date_start = date_start.replace('Z', '+00:00') if 'Z' in date_start else date_start
                        date_start_obj = datetime.fromisoformat(date_start.replace('Z', ''))
                    else:
                        date_start_obj = date_start
                    queryset = queryset.filter(created_at__gte=date_start_obj)
                except Exception as e:
                    print(f"Erreur parsing date_start: {e}")
            
            if date_end:
                try:
                    if isinstance(date_end, str):
                        date_end = date_end.replace('Z', '+00:00') if 'Z' in date_end else date_end
                        date_end_obj = datetime.fromisoformat(date_end.replace('Z', ''))
                    else:
                        date_end_obj = date_end
                    queryset = queryset.filter(created_at__lte=date_end_obj)
                except Exception as e:
                    print(f"Erreur parsing date_end: {e}")
            
            # Filtrer par statut
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            # Filtrer par fournisseur
            if supplier_id:
                queryset = queryset.filter(supplier_id=supplier_id)
            
            # Vérifier qu'il y a des bons de commande
            count = queryset.count()
            if count == 0:
                return Response(
                    {'error': 'Aucun bon de commande trouvé avec les filtres spécifiés'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Limiter le nombre de bons (sécurité)
            queryset = queryset[:500]
            
            # Générer le PDF
            from .services.report_generator_weasy import generate_purchase_orders_report_pdf
            pdf_buffer = generate_purchase_orders_report_pdf(
                queryset,
                request.user,
                date_start_obj,
                date_end_obj
            )
            
            # Créer la réponse HTTP
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            
            filename = f"rapport-bons-commande-{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(response.content)
            
            return response
            
        except ImportError as e:
            print(f"ImportError génération PDF bons de commande: {e}")
            traceback.print_exc()
            return Response(
                {'error': 'Service de génération PDF non disponible', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            print(f"Erreur génération PDF bons de commande: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du rapport: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class InvoiceViewSet(OrganizationFilterMixin, viewsets.ModelViewSet):
    """ViewSet pour les factures"""
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    required_module = Modules.INVOICES
    organization_field = 'organization'  # Use direct organization field
    filterset_fields = ['status', 'client', 'created_by']
    search_fields = ['invoice_number', 'title', 'description']
    ordering_fields = ['created_at', 'total_amount', 'due_date']
    ordering = ['-created_at']
    
    def get_queryset(self):
        # First apply organization filter
        queryset = super().get_queryset()
        
        # Filtre par dates
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        # Filtre pour factures en retard
        overdue = self.request.query_params.get('overdue')
        if overdue == 'true':
            queryset = queryset.filter(
                status='sent',
                due_date__lt=timezone.now().date()
            )
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
        """Ajouter un item à la facture"""
        invoice = self.get_object()
        
        if not invoice.can_be_edited():
            return Response(
                {'error': 'Invoice cannot be edited in current status'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = InvoiceItemSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(invoice=invoice)
            invoice.recalculate_totals()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Envoyer une facture par email"""
        invoice = self.get_object()

        # Permettre l'envoi même si la facture est déjà envoyée (pour relances)
        # if invoice.status != 'draft':
        #     return Response(
        #         {'error': 'Seules les factures en brouillon peuvent être envoyées'},
        #         status=status.HTTP_400_BAD_REQUEST
        #     )

        # Email du destinataire (par défaut: client, sinon fourni dans la requête)
        recipient_email = request.data.get('recipient_email') or request.data.get('email')
        template_type = request.data.get('template', 'classic')
        custom_message = request.data.get('custom_message')
        
        # Récupérer la langue depuis la requête (header Accept-Language ou paramètre)
        language = request.data.get('language')
        if not language:
            # Essayer de récupérer depuis les headers HTTP
            accept_language = request.META.get('HTTP_ACCEPT_LANGUAGE', '')
            if 'en' in accept_language.lower():
                language = 'en'
            else:
                language = getattr(request, 'LANGUAGE_CODE', 'fr')

        # Envoyer l'email avec le PDF
        from .services.email_service import InvoiceEmailService

        result = InvoiceEmailService.send_invoice_email(
            invoice=invoice,
            recipient_email=recipient_email,
            template_type=template_type,
            custom_message=custom_message,
            language=language
        )

        if result['success']:
            # La date d'envoi est déjà sauvegardée dans le service email
            # Les factures ne sont plus en brouillon, mais on garde la vérification par sécurité
            if invoice.status == 'draft':
                invoice.status = 'sent'
                invoice.save(update_fields=['status'])

            serializer = self.get_serializer(invoice)
            return Response({
                'invoice': serializer.data,
                'message': result['message']
            })
        else:
            return Response(
                {'error': result['message']},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def mark_paid(self, request, pk=None):
        """Marquer une facture comme payée en créant un enregistrement de paiement"""
        from apps.invoicing.models import Payment
        from datetime import datetime, date
        from decimal import Decimal
        
        try:
            invoice = self.get_object()
            
            # Vérifier que la facture peut être marquée comme payée
            if invoice.status == 'paid':
                return Response(
                    {'error': 'Invoice is already marked as paid'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if invoice.status == 'cancelled':
                return Response(
                    {'error': 'Cancelled invoices cannot be marked as paid'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Récupérer les informations de paiement depuis la requête
            payment_method = request.data.get('payment_method', 'cash')
            payment_date_str = request.data.get('payment_date')
            notes = request.data.get('notes', '')
            reference_number = request.data.get('reference_number', '')
            transaction_id = request.data.get('transaction_id', '')
            bank_name = request.data.get('bank_name', '')
            check_number = request.data.get('check_number', '')
            
            # Convertir payment_date en objet date
            if payment_date_str:
                try:
                    if isinstance(payment_date_str, str):
                        # Gérer les formats ISO avec ou sans timezone
                        payment_date_str = payment_date_str.replace('Z', '+00:00')
                        payment_date = datetime.fromisoformat(payment_date_str).date()
                    elif isinstance(payment_date_str, date):
                        payment_date = payment_date_str
                    else:
                        payment_date = timezone.now().date()
                except (ValueError, AttributeError):
                    payment_date = timezone.now().date()
            else:
                payment_date = timezone.now().date()
            
            # Calculer le montant à payer (solde dû)
            # Utiliser la même méthode que Payment.clean() pour éviter les problèmes de précision
            total_payments = sum(Decimal(str(p.amount)) for p in invoice.payments.all())
            total_amount = Decimal(str(invoice.total_amount))
            balance_due = total_amount - total_payments
            
            # Arrondir à 2 décimales pour éviter les problèmes de précision
            balance_due = balance_due.quantize(Decimal('0.01'))
            
            # Vérifier que le montant est valide
            if balance_due <= 0:
                return Response(
                    {'error': 'Invoice balance is already zero or negative'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Créer le paiement avec validation explicite
            payment = Payment(
                invoice=invoice,
                amount=balance_due,
                payment_date=payment_date,
                payment_method=payment_method,
                reference_number=reference_number,
                transaction_id=transaction_id,
                notes=notes,
                created_by=request.user,
                status='success',  # Statut pour un paiement réussi
                bank_name=bank_name,
                check_number=check_number,
            )
            
            # Valider avant de sauvegarder
            try:
                payment.full_clean()
            except Exception as validation_error:
                # Si la validation échoue à cause de la précision, essayer avec un montant légèrement inférieur
                if 'dépasse le solde dû' in str(validation_error):
                    # Réduire légèrement le montant pour tenir compte de la précision
                    balance_due = balance_due - Decimal('0.01')
                    if balance_due > 0:
                        payment.amount = balance_due
                        payment.full_clean()
                    else:
                        return Response(
                            {'error': f'Payment validation failed: {str(validation_error)}'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                else:
                    return Response(
                        {'error': f'Payment validation failed: {str(validation_error)}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Sauvegarder le paiement
            payment.save()
            
            # Le statut de la facture sera automatiquement mis à jour via update_status_from_payments()
            # dans la méthode save() du Payment
            
            # Rafraîchir l'invoice depuis la base de données pour obtenir le statut mis à jour
            invoice.refresh_from_db()
            
            serializer = self.get_serializer(invoice)
            return Response(serializer.data)
            
        except Exception as e:
            import traceback
            error_traceback = traceback.format_exc()
            print(f"[ERROR] Error in mark_paid: {e}")
            print(error_traceback)
            return Response(
                {'error': f'Error marking invoice as paid: {str(e)}', 'details': error_traceback},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'], url_path='pdf')
    def generate_pdf(self, request, pk=None):
        """Générer un PDF de la facture avec WeasyPrint (HTML/CSS)"""
        try:
            from django.http import HttpResponse

            invoice = self.get_object()
            template_type = request.query_params.get('template', 'classic')

            # Générer le PDF avec WeasyPrint
            from .services.pdf_generator_weasy import generate_invoice_pdf_weasy
            pdf_buffer = generate_invoice_pdf_weasy(invoice, template_type)
            print(f"✓ PDF généré avec WeasyPrint (template: {template_type})")

            # Créer la réponse HTTP avec le PDF
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )

            # Définir les en-têtes pour le téléchargement
            filename = f"facture-{invoice.invoice_number}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(response.content)

            return response

        except ImportError as e:
            return Response(
                {'error': 'PDF generation service not available', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            return Response(
                {'error': f'Error generating PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='receipt')
    def generate_receipt(self, request, pk=None):
        """Générer un reçu thermal de la facture avec WeasyPrint"""
        try:
            from django.http import HttpResponse
            from django.template.loader import render_to_string
            from weasyprint import HTML
            from django.conf import settings
            import base64
            import qrcode
            import json
            from io import BytesIO

            invoice = self.get_object()

            # Utiliser le générateur pour obtenir les données d'organisation
            from .services.pdf_generator_weasy import InvoiceWeasyPDFGenerator
            generator = InvoiceWeasyPDFGenerator()
            
            # Récupérer les données de l'organisation
            org_data = generator._get_organization_data(invoice)
            
            # Forcer le format thermal pour le reçu
            org_data['paper_size'] = 'thermal_80'
            
            # Générer le QR code
            qr_code_base64 = generator._generate_qr_code(invoice)

            # Préparer le contexte pour le template thermal
            context = {
                'invoice': invoice,
                'organization': org_data,
                'logo_base64': generator._get_logo_base64(org_data),
                'qr_code_base64': qr_code_base64,
                'items': invoice.items.all() if hasattr(invoice, 'items') else [],
                'subtotal': getattr(invoice, 'subtotal', 0) or 0,
                'tax_amount': getattr(invoice, 'tax_amount', 0) or 0,
                'total_amount': getattr(invoice, 'total_amount', 0) or 0,
                'issue_date': getattr(invoice, 'issue_date', None) or getattr(invoice, 'created_at', None),
                'due_date': getattr(invoice, 'due_date', None),
                'client': invoice.client if hasattr(invoice, 'client') else None,
                'template_type': 'thermal',
                'brand_color': org_data.get('brand_color', '#2563eb'),
                'paper_size': 'thermal_80',
                'is_receipt': True,  # Indicateur que c'est un reçu
            }

            # Utiliser le template thermal
            template_name = 'invoicing/pdf_templates/invoice_thermal.html'

            # Rendu HTML
            html_string = render_to_string(template_name, context)

            # Générer le PDF avec WeasyPrint
            html = HTML(string=html_string, base_url=settings.BASE_DIR)
            pdf_bytes = html.write_pdf()

            # Créer la réponse HTTP avec le PDF
            response = HttpResponse(
                pdf_bytes,
                content_type='application/pdf'
            )

            # Définir les en-têtes pour le téléchargement
            filename = f"recu-{invoice.invoice_number}.pdf"
            response['Content-Disposition'] = f'inline; filename="{filename}"'
            response['Content-Length'] = len(response.content)

            return response

        except ImportError as e:
            return Response(
                {'error': 'PDF generation service not available', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error generating receipt: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='create')
    def create_invoice(self, request):
        """Create invoice endpoint for frontend compatibility"""
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'], url_path='bulk-pdf-report')
    def generate_bulk_pdf_report(self, request):
        """Générer un rapport PDF pour plusieurs factures (avec filtres)"""
        from django.http import HttpResponse
        from datetime import datetime
        import traceback
        
        try:
            # Récupérer les paramètres de filtrage
            invoice_ids = request.data.get('invoice_ids', [])
            date_start = request.data.get('date_start')
            date_end = request.data.get('date_end')
            status_filter = request.data.get('status')
            client_id = request.data.get('client_id')
            
            # Construire le queryset
            queryset = self.get_queryset()
            
            # Filtrer par IDs si fournis
            if invoice_ids:
                queryset = queryset.filter(id__in=invoice_ids)
            
            # Filtrer par dates
            date_start_obj = None
            date_end_obj = None
            if date_start:
                try:
                    # Gérer différents formats de date
                    # Les dates viennent du frontend au format YYYY-MM-DD (input type="date")
                    if isinstance(date_start, str):
                        from datetime import date as date_class
                        # Si c'est juste une date (YYYY-MM-DD), utiliser date.fromisoformat
                        if 'T' not in date_start and ' ' not in date_start:
                            date_start_obj = date_class.fromisoformat(date_start)
                        else:
                            # Si c'est un datetime, extraire juste la date
                            date_start_str = date_start.split('T')[0].split(' ')[0]
                            date_start_obj = date_class.fromisoformat(date_start_str)
                    elif isinstance(date_start, datetime):
                        date_start_obj = date_start.date()
                    elif hasattr(date_start, 'date'):
                        date_start_obj = date_start.date()
                    else:
                        date_start_obj = date_start
                    
                    if date_start_obj:
                        # Utiliser created_at au lieu de issue_date (qui n'existe pas)
                        queryset = queryset.filter(created_at__date__gte=date_start_obj)
                except Exception as e:
                    print(f"Erreur parsing date_start '{date_start}': {e}")
                    import traceback
                    traceback.print_exc()
            
            if date_end:
                try:
                    if isinstance(date_end, str):
                        from datetime import date as date_class
                        # Si c'est juste une date (YYYY-MM-DD), utiliser date.fromisoformat
                        if 'T' not in date_end and ' ' not in date_end:
                            date_end_obj = date_class.fromisoformat(date_end)
                        else:
                            # Si c'est un datetime, extraire juste la date
                            date_end_str = date_end.split('T')[0].split(' ')[0]
                            date_end_obj = date_class.fromisoformat(date_end_str)
                    elif isinstance(date_end, datetime):
                        date_end_obj = date_end.date()
                    elif hasattr(date_end, 'date'):
                        date_end_obj = date_end.date()
                    else:
                        date_end_obj = date_end
                    
                    if date_end_obj:
                        # Utiliser created_at au lieu de issue_date (qui n'existe pas)
                        queryset = queryset.filter(created_at__date__lte=date_end_obj)
                except Exception as e:
                    print(f"Erreur parsing date_end '{date_end}': {e}")
                    import traceback
                    traceback.print_exc()
            
            # Filtrer par statut
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            
            # Filtrer par client
            if client_id:
                queryset = queryset.filter(client_id=client_id)
            
            # Précharger les relations pour éviter les problèmes de lazy loading
            # IMPORTANT: select_related doit être appelé AVANT la limitation [:500]
            try:
                queryset = queryset.select_related('client', 'created_by')
            except Exception as e:
                print(f"Erreur select_related: {e}")
            
            # Vérifier qu'il y a des factures
            count = queryset.count()
            if count == 0:
                return Response(
                    {'error': 'Aucune facture trouvée avec les filtres spécifiés'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Limiter le nombre de factures (sécurité) - après select_related
            queryset = queryset[:500]
            
            # Générer le PDF
            from .services.report_generator_weasy import generate_invoices_report_pdf
            pdf_buffer = generate_invoices_report_pdf(
                queryset, 
                request.user,
                date_start_obj,
                date_end_obj
            )
            
            # Créer la réponse HTTP
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            
            filename = f"rapport-factures-{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(response.content)
            
            return response
            
        except ImportError as e:
            print(f"ImportError génération PDF factures: {e}")
            traceback.print_exc()
            return Response(
                {'error': 'Service de génération PDF non disponible', 'details': str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        except Exception as e:
            print(f"Erreur génération PDF factures: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Erreur lors de la génération du rapport: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def cancel_with_credit_note(self, request, pk=None):
        """
        Cancel invoice and generate credit note
        Permissions: Managers + Administrators only
        """
        from django.utils import timezone

        invoice = self.get_object()

        # Check permissions - only managers and admins can cancel
        user_role = getattr(request.user, 'role', None)
        if user_role not in ['manager', 'admin']:
            return Response(
                {'error': 'Only managers and administrators can cancel invoices'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Validate cancellation
        if invoice.status == 'cancelled':
            return Response(
                {'error': 'Invoice is already cancelled'},
                status=status.HTTP_400_BAD_REQUEST
            )

        force_cancel = request.data.get('force_cancel', False)
        if invoice.status == 'paid' and not force_cancel:
            return Response(
                {'error': 'Paid invoices cannot be cancelled without force_cancel flag'},
                status=status.HTTP_400_BAD_REQUEST
            )

        cancellation_reason = request.data.get('reason', '')
        if not cancellation_reason or not cancellation_reason.strip():
            return Response(
                {'error': 'Cancellation reason is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Create credit note (negative invoice)
            credit_note = Invoice.objects.create(
                organization=invoice.organization,
                invoice_type='credit_note',
                status='sent',
                client=invoice.client,
                created_by=request.user,
                subtotal=-invoice.subtotal,
                tax_amount=-invoice.tax_amount,
                total_amount=-invoice.total_amount,
                original_invoice=invoice,
                title=f"Avoir - {invoice.invoice_number}",
                description=f"Annulation de {invoice.invoice_number}: {cancellation_reason}",
                currency=invoice.currency,
                billing_address=invoice.billing_address,
                payment_terms=invoice.payment_terms
            )

            # Copy items with negative unit prices
            for item in invoice.items.all():
                InvoiceItem.objects.create(
                    invoice=credit_note,
                    product=item.product,
                    service_code=item.service_code,
                    product_reference=item.product_reference,
                    description=item.description,
                    detailed_description=item.detailed_description,
                    quantity=item.quantity,
                    unit_price=-item.unit_price,
                    total_price=-item.total_price,
                    unit_of_measure=item.unit_of_measure,
                    discount_percent=item.discount_percent,
                    tax_rate=item.tax_rate,
                    notes=item.notes
                )

            # Update original invoice
            invoice.status = 'cancelled'
            invoice.cancellation_reason = cancellation_reason
            invoice.cancelled_at = timezone.now()
            invoice.cancelled_by = request.user
            invoice.save()

            # Create audit trail log (if activity logger exists)
            try:
                from apps.analytics.activity_logger import log_activity
                log_activity(
                    user=request.user,
                    action='invoice_cancelled',
                    model_name='Invoice',
                    object_id=str(invoice.id),
                    details={
                        'invoice_number': invoice.invoice_number,
                        'credit_note_number': credit_note.invoice_number,
                        'reason': cancellation_reason,
                        'original_amount': float(invoice.total_amount)
                    }
                )
            except ImportError:
                # Activity logger not available
                pass

            # Serialize both invoices
            serializer = self.get_serializer(invoice)
            credit_note_serializer = self.get_serializer(credit_note)

            return Response({
                'message': 'Invoice cancelled successfully',
                'invoice': serializer.data,
                'credit_note': credit_note_serializer.data
            })

        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Error cancelling invoice: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DashboardStatsView(APIView):
    """Vue pour les statistiques du tableau de bord"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from apps.core.modules import get_user_accessible_modules
        from datetime import datetime

        # Get user's organization
        user = request.user
        organization = getattr(user, 'organization', None)

        if not organization:
            return Response({
                'enabled_modules': [],
                'error': 'No organization assigned to user'
            }, status=status.HTTP_403_FORBIDDEN)

        # Récupérer les modules accessibles par l'utilisateur
        user_modules = get_user_accessible_modules(user)

        # Initialiser les stats
        stats = {
            'enabled_modules': user_modules,
        }

        # Stats Fournisseurs (si module actif) - FILTERED BY ORGANIZATION
        if Modules.SUPPLIERS in user_modules:
            suppliers = Supplier.objects.filter(organization=organization)
            stats['total_suppliers'] = suppliers.count()
            stats['active_suppliers'] = suppliers.filter(status='active').count()
            stats['inactive_suppliers'] = suppliers.filter(status='inactive').count()
            stats['suppliers_by_rating'] = {
                '5_stars': suppliers.filter(rating=5).count(),
                '4_stars': suppliers.filter(rating=4).count(),
                '3_stars': suppliers.filter(rating=3).count(),
                'below_3': suppliers.filter(rating__lt=3).count(),
            }

        # Stats Bons de commande (si module actif) - FILTERED BY ORGANIZATION
        if Modules.PURCHASE_ORDERS in user_modules:
            purchase_orders = PurchaseOrder.objects.filter(created_by__organization=organization)
            stats['total_purchase_orders'] = purchase_orders.count()
            stats['pending_purchase_orders'] = purchase_orders.filter(
                status__in=['draft', 'pending']
            ).count()
            stats['approved_purchase_orders'] = purchase_orders.filter(status='approved').count()
            stats['received_purchase_orders'] = purchase_orders.filter(status='received').count()
            stats['total_expenses'] = purchase_orders.filter(
                status__in=['approved', 'sent', 'received']
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

        # Stats Factures (si module actif) - FILTERED BY ORGANIZATION
        if Modules.INVOICES in user_modules:
            invoices = Invoice.objects.filter(created_by__organization=organization)
            stats['total_invoices'] = invoices.count()
            stats['unpaid_invoices'] = invoices.filter(status='sent').count()
            stats['paid_invoices'] = invoices.filter(status='paid').count()
            stats['overdue_invoices'] = invoices.filter(status='overdue').count()
            stats['draft_invoices'] = invoices.filter(status='draft').count()
            stats['total_revenue'] = invoices.filter(
                status='paid'
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            stats['pending_revenue'] = invoices.filter(
                status='sent'
            ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

        # Stats Produits (si module actif) - FILTERED BY ORGANIZATION
        if Modules.PRODUCTS in user_modules:
            products = Product.objects.filter(organization=organization)
            stats['total_products'] = products.count()
            stats['active_products'] = products.filter(is_active=True).count()
            stats['low_stock_products'] = products.filter(
                product_type='physical',
                stock_quantity__lte=F('low_stock_threshold')
            ).count()
            stats['out_of_stock_products'] = products.filter(
                product_type='physical',
                stock_quantity=0
            ).count()
            stats['total_stock_value'] = products.filter(
                product_type='physical'
            ).aggregate(
                total=Sum(F('stock_quantity') * F('cost_price'))
            )['total'] or 0

        # Stats Clients (si module actif) - FILTERED BY ORGANIZATION
        if Modules.CLIENTS in user_modules:
            clients = Client.objects.filter(organization=organization)
            stats['total_clients'] = clients.count()
            stats['active_clients'] = clients.filter(is_active=True).count()
            # Clients avec au moins une facture payée dans les 3 derniers mois
            three_months_ago = timezone.now() - timedelta(days=90)
            stats['recent_active_clients'] = clients.filter(
                invoices__status='paid',
                invoices__updated_at__gte=three_months_ago
            ).distinct().count()

        # Stats de performance globale
        if Modules.INVOICES in user_modules and Modules.PURCHASE_ORDERS in user_modules:
            revenue = stats.get('total_revenue', 0)
            expenses = stats.get('total_expenses', 0)
            stats['net_profit'] = revenue - expenses
            stats['profit_margin'] = ((revenue - expenses) / revenue * 100) if revenue > 0 else 0

        # Tendances (30 derniers jours)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        sixty_days_ago = timezone.now() - timedelta(days=60)

        if Modules.INVOICES in user_modules:
            invoices = Invoice.objects.filter(created_by__organization=organization)
            recent_invoices_count = invoices.filter(created_at__gte=thirty_days_ago).count()
            previous_invoices_count = invoices.filter(
                created_at__gte=sixty_days_ago,
                created_at__lt=thirty_days_ago
            ).count()
            stats['invoices_trend'] = {
                'current': recent_invoices_count,
                'previous': previous_invoices_count,
                'percent_change': ((recent_invoices_count - previous_invoices_count) / previous_invoices_count * 100) if previous_invoices_count > 0 else 0
            }

        if Modules.PURCHASE_ORDERS in user_modules:
            purchase_orders = PurchaseOrder.objects.filter(created_by__organization=organization)
            recent_po_count = purchase_orders.filter(created_at__gte=thirty_days_ago).count()
            previous_po_count = purchase_orders.filter(
                created_at__gte=sixty_days_ago,
                created_at__lt=thirty_days_ago
            ).count()
            stats['purchase_orders_trend'] = {
                'current': recent_po_count,
                'previous': previous_po_count,
                'percent_change': ((recent_po_count - previous_po_count) / previous_po_count * 100) if previous_po_count > 0 else 0
            }

        serializer = DashboardStatsSerializer(stats)
        return Response(serializer.data)


class RecentActivityView(APIView):
    """Vue pour l'activité récente"""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            # Get user's organization
            user = request.user
            organization = getattr(user, 'organization', None)

            if not organization:
                return Response({
                    'recent_suppliers': [],
                    'recent_purchase_orders': [],
                    'recent_invoices': [],
                    'error': 'No organization assigned to user'
                }, status=status.HTTP_403_FORBIDDEN)

            # Derniers 7 jours
            since = timezone.now() - timedelta(days=7)

            recent_data = {
                'recent_suppliers': [],
                'recent_purchase_orders': [],
                'recent_invoices': [],
            }

            # Essayer de récupérer les données récentes - FILTERED BY ORGANIZATION
            try:
                recent_suppliers = Supplier.objects.filter(
                    organization=organization,
                    created_at__gte=since
                ).order_by('-created_at')[:5]
                recent_data['recent_suppliers'] = SupplierSerializer(recent_suppliers, many=True).data
            except Exception as e:
                print(f"Erreur suppliers: {e}")
                recent_data['recent_suppliers'] = []

            try:
                recent_orders = PurchaseOrder.objects.filter(
                    created_by__organization=organization,
                    created_at__gte=since
                ).order_by('-created_at')[:5]
                recent_data['recent_purchase_orders'] = PurchaseOrderSerializer(recent_orders, many=True).data
            except Exception as e:
                print(f"Erreur purchase orders: {e}")
                recent_data['recent_purchase_orders'] = []

            try:
                recent_invoices = Invoice.objects.filter(
                    created_by__organization=organization,
                    created_at__gte=since
                ).order_by('-created_at')[:5]
                recent_data['recent_invoices'] = InvoiceSerializer(recent_invoices, many=True).data
            except Exception as e:
                print(f"Erreur invoices: {e}")
                recent_data['recent_invoices'] = []

            return Response(recent_data)
        except Exception as e:
            print(f"Erreur générale RecentActivityView: {e}")
            return Response({
                'recent_suppliers': [],
                'recent_purchase_orders': [],
                'recent_invoices': [],
                'error': str(e)
            }, status=500)


class AIConversationsView(APIView):
    """Vue temporaire pour les conversations IA"""
    permission_classes = [permissions.AllowAny]  # Temporaire pour le développement
    
    def get(self, request):
        # Retourner des données vides pour l'instant
        return Response([])


class AIQuickActionsView(APIView):
    """Vue temporaire pour les actions rapides IA"""
    permission_classes = [permissions.AllowAny]  # Temporaire pour le développement
    
    def get(self, request):
        # Actions rapides temporaires
        actions = [
            {
                'id': 'create_supplier',
                'title': 'Créer un fournisseur',
                'description': 'Ajouter un nouveau fournisseur',
                'icon': 'person_add',
                'url': '/suppliers/create/'
            },
            {
                'id': 'create_purchase_order',
                'title': 'Nouveau bon de commande',
                'description': 'Créer un bon de commande',
                'icon': 'shopping_cart',
                'url': '/purchase-orders/create/'
            },
            {
                'id': 'create_invoice',
                'title': 'Nouvelle facture',
                'description': 'Créer une facture',
                'icon': 'receipt',
                'url': '/invoicing/create/'
            },
            {
                'id': 'view_dashboard',
                'title': 'Tableau de bord',
                'description': 'Voir les statistiques',
                'icon': 'dashboard',
                'url': '/'
            }
        ]
        return Response(actions)


class ProductCategoryViewSet(viewsets.ModelViewSet):
    """ViewSet pour les catégories de produits"""
    queryset = ProductCategory.objects.all()
    serializer_class = ProductCategorySerializer
    permission_classes = [permissions.IsAuthenticated]
    # required_module removed - categories should be accessible regardless of module activation
    filterset_fields = ['is_active', 'parent']
    search_fields = ['name', 'slug', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtre par organization de l'utilisateur
        if self.request.user.is_authenticated and hasattr(self.request.user, 'organization') and self.request.user.organization:
            queryset = queryset.filter(organization=self.request.user.organization)

        return queryset


class WarehouseViewSet(viewsets.ModelViewSet):
    """ViewSet pour les entrepôts"""
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    permission_classes = [permissions.IsAuthenticated]
    # required_module removed - warehouses should be accessible regardless of module activation
    filterset_fields = ['is_active', 'city', 'province']
    search_fields = ['name', 'code', 'address', 'city']
    ordering_fields = ['name', 'code', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtre par organization de l'utilisateur
        if self.request.user.is_authenticated and hasattr(self.request.user, 'organization') and self.request.user.organization:
            queryset = queryset.filter(organization=self.request.user.organization)

        return queryset

class StockMovementCancelView(APIView):
    """
    DELETE /api/stock-movements/<uuid>/cancel/
    Annule (supprime) un mouvement de stock et remet le stock a l'etat anterieur.
    Reserve aux administrateurs et managers.
    """
    permission_classes = [permissions.IsAuthenticated]

    def delete(self, request, movement_id):
        from apps.invoicing.models import StockMovement

        user = request.user
        if not (user.is_staff or getattr(user, 'profile_type', None) in ('admin', 'manager')):
            return Response(
                {'error': 'Permission refusee. Seuls les administrateurs peuvent annuler des mouvements.'},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            movement = StockMovement.objects.select_related('product').get(
                id=movement_id,
                product__organization=user.organization
            )
        except StockMovement.DoesNotExist:
            return Response({'error': 'Mouvement introuvable'}, status=status.HTTP_404_NOT_FOUND)

        product = movement.product
        product.stock_quantity -= movement.quantity
        product.save(update_fields=['stock_quantity'])

        movement_info = {
            'id': str(movement.id),
            'movement_type': movement.movement_type,
            'quantity': movement.quantity,
        }

        movement.delete()

        return Response({
            'message': 'Mouvement annule et stock mis a jour',
            'cancelled_movement': movement_info,
            'product_stock': product.stock_quantity,
        }, status=status.HTTP_200_OK)
