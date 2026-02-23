"""
Export au format ERPNext pour la comptabilité.

Formats disponibles via ?format= :
  - csv       : CSV simple (Excel-friendly, séparateur virgule, UTF-8 BOM)
  - xlsx      : Excel (.xlsx, openpyxl)
  - erpnext   : CSV avec les 19 lignes d'en-tête ERPNext (pour import direct dans ERPNext)

Endpoints (admin/manager uniquement) :
  GET /api/erpnext/sales-invoices/?date_from=&date_to=&format=csv|xlsx|erpnext
  GET /api/erpnext/customers/?date_from=&date_to=&format=
  GET /api/erpnext/items/?format=
  GET /api/erpnext/payments/?date_from=&date_to=&format=
  GET /api/erpnext/purchase-orders/?date_from=&date_to=&format=
"""
print("DEBUG: Loading erpnext_export_views.py")

import csv
from datetime import date, datetime
from io import StringIO, BytesIO

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status as drf_status


# ──────────────────────────────────────────────
#  Helpers
# ──────────────────────────────────────────────

def _require_admin(user):
    return user.is_authenticated and user.role in ('admin', 'manager')


def _parse_date(val, default):
    if not val:
        return default
    try:
        return datetime.strptime(val, '%Y-%m-%d').date()
    except ValueError:
        return default


def _get_org(user):
    return getattr(user, 'organization', None)


def _simple_csv_response(filename, columns, rows):
    """CSV simple avec BOM UTF-8 pour Excel."""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([c['label'] for c in columns])
    for row in rows:
        writer.writerow(row)
    content = '\ufeff' + output.getvalue()   # BOM UTF-8 pour Excel
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


def _xlsx_response(filename, columns, rows, sheet_title='Export'):
    """Export Excel .xlsx avec openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # En-tête
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col['label'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(15, len(col['label']) + 4)

    # Données
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def _erpnext_csv_response(filename, doctype, columns, rows, child_columns=None, child_doctype=None, prepend_empty=True):
    """
    CSV au format natif ERPNext (avec les 19 lignes de méta-données).
    - columns : liste de dicts {label, name, required, type, info}
    - child_columns : colonnes de la table enfant (optionnel)
    - child_doctype : nom de la table enfant ERPNext (ex: 'Sales Invoice Item')
    """
    INSTRUCTIONS = [
        "Veuillez ne pas modifier les sections du modèle.",
        "La première colonne de données doit être vide.",
        'Si vous chargez de nouveaux enregistrements, laissez la colonne "nom" (ID) vide.',
        'Si vous téléchargez de nouveaux rapports, "Nommer Séries" devient obligatoire, si présent.',
        "Seuls les champs obligatoires sont nécessaires pour les nouveaux enregistrements. Vous pouvez supprimer des colonnes non obligatoires si vous le souhaitez.",
        "Pour la mise à jour, vous pouvez mettre à jour uniquement une sélection colonnes.",
        "Vous pouvez seulement charger jusqu'à 5000 enregistrement en une seule fois. (peut-être moins dans certains cas)",
    ]

    all_columns = list(columns)
    has_child = child_columns and child_doctype

    if has_child:
        all_columns += [{'label': '', 'name': '~', 'required': '', 'type': '', 'info': ''}]
        all_columns += child_columns

    output = StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)

    # Lignes 1-2
    writer.writerow(["Modèle d'importation de données"])
    writer.writerow(["Table:", doctype])
    writer.writerow([])
    writer.writerow([])
    # Ligne 5
    writer.writerow(["Remarques:"])
    for inst in INSTRUCTIONS:
        writer.writerow([inst])
    writer.writerow([])

    # Ligne 13 : DocType header
    doctype_row = ["DocType:", doctype]
    doctype_row += [''] * (len(columns) - 1)
    if has_child:
        doctype_row += ['~', child_doctype]
        doctype_row += [''] * (len(child_columns) - 2)
    writer.writerow(doctype_row)

    # Lignes 14-18 : méta-données colonnes
    writer.writerow(["Étiquettes de Colonne :"] + [c['label'] for c in all_columns])
    writer.writerow(["Nom de la Colonne:"] + [c['name'] for c in all_columns])
    writer.writerow(["Obligatoire :"] + [c.get('required', 'No') for c in all_columns])
    writer.writerow(["Type :"] + [c.get('type', 'Data') for c in all_columns])
    writer.writerow(["Info :"] + [c.get('info', '') for c in all_columns])

    # Ligne 19
    writer.writerow(["Commencez à entrer les données ci-dessous"])

    # Données (ERPNext attend une colonne vide en début de chaque ligne de données)
    for row in rows:
        if prepend_empty:
            writer.writerow([''] + list(row))
        else:
            writer.writerow(list(row))

    content = '\ufeff' + output.getvalue()
    response = HttpResponse(content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}_erpnext.csv"'
    return response


def _dispatch_response(fmt, filename, columns, rows, doctype, child_columns=None, child_doctype=None, sheet_title='Export'):
    """Dispatch vers le bon format selon ?format=."""
    if fmt == 'xlsx':
        return _xlsx_response(filename, columns, rows, sheet_title=sheet_title)
    elif fmt == 'erpnext':
        return _erpnext_csv_response(filename, doctype, columns, rows, child_columns, child_doctype)
    else:
        return _simple_csv_response(filename, columns, rows)


# ──────────────────────────────────────────────
#  Définitions des colonnes ERPNext par DocType
# ──────────────────────────────────────────────

CUSTOMER_COLUMNS = [
    {'label': 'ID',              'name': 'name',          'required': 'Yes', 'type': 'Data',   'info': ''},
    {'label': 'Nom du client',   'name': 'customer_name', 'required': 'Yes', 'type': 'Data',   'info': ''},
    {'label': 'Type de client',  'name': 'customer_type', 'required': 'Yes', 'type': 'Select', 'info': 'Un des: Company, Individual'},
    {'label': 'Email',           'name': 'email_id',      'required': 'No',  'type': 'Data',   'info': ''},
    {'label': 'Mobile',          'name': 'mobile_no',     'required': 'No',  'type': 'Data',   'info': ''},
    {'label': 'Territoire',      'name': 'territory',     'required': 'No',  'type': 'Link',   'info': 'Valid Territory'},
    {'label': 'Groupe de clients','name':'customer_group','required': 'No',  'type': 'Link',   'info': 'Valid Customer Group'},
    {'label': 'Devise',          'name': 'default_currency','required':'No', 'type': 'Link',   'info': 'Valid Currency'},
]

ITEM_COLUMNS = [
    {'label': 'ID',                'name': 'name',             'required': 'Yes', 'type': 'Data',   'info': ''},
    {'label': 'Code article',      'name': 'item_code',        'required': 'Yes', 'type': 'Data',   'info': ''},
    {'label': 'Nom de l\'article', 'name': 'item_name',        'required': 'Yes', 'type': 'Data',   'info': ''},
    {'label': 'Groupe d\'article', 'name': 'item_group',       'required': 'Yes', 'type': 'Link',   'info': 'Valid Item Group'},
    {'label': 'Description',       'name': 'description',      'required': 'No',  'type': 'Text',   'info': ''},
    {'label': 'Unité de mesure',   'name': 'stock_uom',        'required': 'No',  'type': 'Link',   'info': 'Valid UOM'},
    {'label': 'Est un article de stock', 'name': 'is_stock_item', 'required': 'No', 'type': 'Check', 'info': '0 or 1'},
    {'label': 'Prix de vente std', 'name': 'standard_rate',    'required': 'No',  'type': 'Currency','info': ''},
    {'label': 'Stock actuel',      'name': 'opening_stock',    'required': 'No',  'type': 'Float',  'info': ''},
    {'label': 'Valuation Rate',    'name': 'valuation_rate',   'required': 'No',  'type': 'Currency','info': ''},
    {'label': 'Fournisseur',       'name': 'default_supplier', 'required': 'No',  'type': 'Link',   'info': 'Valid Supplier'},
    {'label': 'Désactivé',         'name': 'disabled',         'required': 'No',  'type': 'Check',  'info': '0 or 1'},
]

SALES_INVOICE_COLUMNS = [
    {'label': 'ID',               'name': 'name',             'required': 'Yes', 'type': 'Data',   'info': ''},
    {'label': 'Client',           'name': 'customer',         'required': 'Yes', 'type': 'Link',   'info': 'Valid Customer'},
    {'label': 'Date de comptabilisation', 'name': 'posting_date', 'required': 'Yes', 'type': 'Date', 'info': 'dd/mm/yyyy'},
    {'label': 'Date d\'échéance', 'name': 'due_date',         'required': 'No',  'type': 'Date',   'info': 'dd/mm/yyyy'},
    {'label': 'Société',          'name': 'company',          'required': 'Yes', 'type': 'Link',   'info': 'Valid Company'},
    {'label': 'Devise',           'name': 'currency',         'required': 'No',  'type': 'Link',   'info': 'Valid Currency'},
    {'label': 'Mode de paiement', 'name': 'payment_terms_template', 'required': 'No', 'type': 'Link', 'info': ''},
    {'label': 'Numéro de pièce',  'name': 'po_no',            'required': 'No',  'type': 'Data',   'info': ''},
    {'label': 'Remarques',        'name': 'remarks',          'required': 'No',  'type': 'Text',   'info': ''},
]

SALES_INVOICE_ITEM_COLUMNS = [
    {'label': 'ID',                'name': 'name',        'required': 'Yes', 'type': 'Data',     'info': ''},
    {'label': 'Code article',      'name': 'item_code',   'required': 'Yes', 'type': 'Link',     'info': 'Valid Item'},
    {'label': 'Nom de l\'article', 'name': 'item_name',   'required': 'No',  'type': 'Data',     'info': ''},
    {'label': 'Quantité',          'name': 'qty',         'required': 'Yes', 'type': 'Float',    'info': ''},
    {'label': 'Taux (Prix unit.)', 'name': 'rate',        'required': 'Yes', 'type': 'Currency', 'info': ''},
    {'label': 'Total',             'name': 'amount',      'required': 'No',  'type': 'Currency', 'info': ''},
    {'label': 'Unité de mesure',   'name': 'uom',         'required': 'No',  'type': 'Link',     'info': 'Valid UOM'},
    {'label': 'Description',       'name': 'description', 'required': 'No',  'type': 'Text',     'info': ''},
]

PAYMENT_COLUMNS = [
    {'label': 'ID',               'name': 'name',               'required': 'Yes', 'type': 'Data',    'info': ''},
    {'label': 'Type de paiement', 'name': 'payment_type',       'required': 'Yes', 'type': 'Select',  'info': 'Un des: Receive, Pay, Internal Transfer'},
    {'label': 'Date de comptabilisation', 'name': 'posting_date','required': 'Yes', 'type': 'Date',   'info': 'dd/mm/yyyy'},
    {'label': 'Société',          'name': 'company',            'required': 'Yes', 'type': 'Link',    'info': 'Valid Company'},
    {'label': 'Type de tiers',    'name': 'party_type',         'required': 'Yes', 'type': 'Select',  'info': 'Un des: Customer, Supplier'},
    {'label': 'Tiers',            'name': 'party',              'required': 'Yes', 'type': 'Dynamic Link', 'info': ''},
    {'label': 'Montant payé',     'name': 'paid_amount',        'required': 'Yes', 'type': 'Currency','info': ''},
    {'label': 'Devise payée',     'name': 'paid_from_account_currency', 'required': 'No', 'type': 'Link', 'info': 'Valid Currency'},
    {'label': 'Mode de paiement', 'name': 'mode_of_payment',   'required': 'No',  'type': 'Link',    'info': 'Valid Mode of Payment'},
    {'label': 'Référence',        'name': 'reference_no',       'required': 'No',  'type': 'Data',    'info': ''},
    {'label': 'Date de référence','name': 'reference_date',     'required': 'No',  'type': 'Date',    'info': 'dd/mm/yyyy'},
    {'label': 'Remarques',        'name': 'remarks',            'required': 'No',  'type': 'Text',    'info': ''},
]

PURCHASE_ORDER_COLUMNS = [
    {'label': 'ID',             'name': 'name',          'required': 'Yes', 'type': 'Data',  'info': ''},
    {'label': 'Fournisseur',    'name': 'supplier',      'required': 'Yes', 'type': 'Link',  'info': 'Valid Supplier'},
    {'label': 'Date de transaction', 'name': 'transaction_date', 'required': 'Yes', 'type': 'Date', 'info': 'dd/mm/yyyy'},
    {'label': 'Société',        'name': 'company',       'required': 'Yes', 'type': 'Link',  'info': 'Valid Company'},
    {'label': 'Devise',         'name': 'currency',      'required': 'No',  'type': 'Link',  'info': 'Valid Currency'},
    {'label': 'Titre',          'name': 'title',         'required': 'No',  'type': 'Data',  'info': ''},
]

PURCHASE_ORDER_ITEM_COLUMNS = [
    {'label': 'ID',                'name': 'name',        'required': 'Yes', 'type': 'Data',     'info': ''},
    {'label': 'Code article',      'name': 'item_code',   'required': 'Yes', 'type': 'Link',     'info': 'Valid Item'},
    {'label': 'Nom de l\'article', 'name': 'item_name',   'required': 'No',  'type': 'Data',     'info': ''},
    {'label': 'Quantité',          'name': 'qty',         'required': 'Yes', 'type': 'Float',    'info': ''},
    {'label': 'Taux',              'name': 'rate',        'required': 'Yes', 'type': 'Currency', 'info': ''},
    {'label': 'Total',             'name': 'amount',      'required': 'No',  'type': 'Currency', 'info': ''},
    {'label': 'Unité de mesure',   'name': 'uom',         'required': 'No',  'type': 'Link',     'info': 'Valid UOM'},
    {'label': 'Date de livraison', 'name': 'schedule_date','required': 'No', 'type': 'Date',     'info': 'dd/mm/yyyy'},
]


# ──────────────────────────────────────────────
#  Endpoints
# ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_customers(request):
    """Export clients → ERPNext Customer"""
    if not _require_admin(request.user):
        return Response({'detail': 'Accès réservé aux administrateurs.'}, status=drf_status.HTTP_403_FORBIDDEN)

    from apps.accounts.models import Client

    org = _get_org(request.user)
    if not org:
        return Response({'detail': 'Organisation introuvable.'}, status=drf_status.HTTP_400_BAD_REQUEST)

    fmt = request.GET.get('format', 'csv')
    date_from = _parse_date(request.GET.get('date_from'), None)
    date_to = _parse_date(request.GET.get('date_to'), None)

    qs = Client.objects.filter(organization=org).order_by('name')
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    rows = []
    for c in qs:
        erpnext_type = 'Individual' if c.client_type in ('patient', 'both') else 'Company'
        rows.append([
            '',                     # ID (vide = nouveau)
            c.name,                 # customer_name
            erpnext_type,           # customer_type
            c.email or '',          # email_id
            c.phone or '',          # mobile_no
            '',                     # territory
            '',                     # customer_group
            'XAF',                  # default_currency
        ])

    return _dispatch_response(fmt, f'erpnext_clients_{date.today()}',
                               CUSTOMER_COLUMNS, rows,
                               doctype='Customer',
                               sheet_title='Clients')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_items(request):
    """Export produits → ERPNext Item"""
    if not _require_admin(request.user):
        return Response({'detail': 'Accès réservé aux administrateurs.'}, status=drf_status.HTTP_403_FORBIDDEN)

    from apps.invoicing.models import Product

    org = _get_org(request.user)
    if not org:
        return Response({'detail': 'Organisation introuvable.'}, status=drf_status.HTTP_400_BAD_REQUEST)

    fmt = request.GET.get('format', 'csv')
    qs = Product.objects.filter(organization=org, is_active=True).select_related('category', 'supplier')

    rows = []
    for p in qs:
        is_stock = 1 if p.product_type == 'physical' else 0
        item_group = p.category.name if p.category else 'All Item Groups'
        rows.append([
            '',                                         # name (ID)
            p.reference or p.name[:20],                 # item_code
            p.name,                                     # item_name
            item_group,                                 # item_group
            p.description or '',                        # description
            _map_uom(p.sell_unit),                      # stock_uom
            is_stock,                                   # is_stock_item
            float(p.price),                             # standard_rate
            float(p.stock_quantity) if is_stock else 0, # opening_stock
            float(p.cost_price) if p.cost_price else 0, # valuation_rate
            p.supplier.name if p.supplier else '',       # default_supplier
            0,                                          # disabled
        ])

    return _dispatch_response(fmt, f'erpnext_articles_{date.today()}',
                               ITEM_COLUMNS, rows,
                               doctype='Item',
                               sheet_title='Articles')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_sales_invoices(request):
    """Export factures → ERPNext Sales Invoice (avec lignes article)"""
    print("DEBUG: Entered export_sales_invoices")
    if not _require_admin(request.user):
        print("DEBUG: Not an admin")
        return Response({'detail': 'Accès réservé aux administrateurs.'}, status=drf_status.HTTP_403_FORBIDDEN)

    from apps.invoicing.models import Invoice

    org = _get_org(request.user)
    print(f"DEBUG: Organization: {org}")
    if not org:
        return Response({'detail': 'Organisation introuvable.'}, status=drf_status.HTTP_400_BAD_REQUEST)

    fmt = request.GET.get('format', 'csv')
    date_from = _parse_date(request.GET.get('date_from'), date(date.today().year, 1, 1))
    date_to = _parse_date(request.GET.get('date_to'), date.today())

    invoices = Invoice.objects.filter(
        organization=org,
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).exclude(status='cancelled').select_related('client').prefetch_related('items__product').order_by('created_at')

    company = org.name

    if fmt == 'erpnext':
        # Format ERPNext : CSV avec les 19 lignes de méta-données + table enfant items
        n_main = len(SALES_INVOICE_COLUMNS)
        n_child = len(SALES_INVOICE_ITEM_COLUMNS)

        erpnext_rows = []
        for inv in invoices:
            client_name = inv.client.name if inv.client else ''
            inv_date = inv.created_at.date().strftime('%d/%m/%Y')
            due_date = inv.due_date.strftime('%d/%m/%Y') if inv.due_date else inv_date
            items = list(inv.items.all())

            # Valeurs des colonnes principales
            main_vals = ['', client_name, inv_date, due_date, company, 'XAF', '', inv.invoice_number, inv.invoice_type]

            if not items:
                # Ligne sans items : remplir les colonnes child avec vide
                erpnext_rows.append(main_vals + [''] * (1 + n_child))  # ~ + child cols
            else:
                for i, item in enumerate(items):
                    item_code = (item.product.reference or item.product.name[:20]) if item.product else ''
                    item_name = item.product.name if item.product else (item.description or '')
                    uom = _map_uom(item.product.sell_unit if item.product else 'piece')
                    child_vals = [
                        '',          # ID item (vide = nouveau)
                        item_code,   # item_code
                        item_name,   # item_name
                        float(item.quantity),   # qty
                        float(item.unit_price), # rate
                        float(item.total_price),# amount
                        uom,                    # uom
                        item.description or '', # description
                    ]
                    if i == 0:
                        erpnext_rows.append(main_vals + [''] + child_vals)
                    else:
                        # Lignes suivantes : colonnes principales vides
                        erpnext_rows.append([''] * n_main + [''] + child_vals)

        return _erpnext_csv_response(
            f'erpnext_factures_{date_from}_{date_to}',
            'Sales Invoice',
            SALES_INVOICE_COLUMNS,
            erpnext_rows,
            child_columns=SALES_INVOICE_ITEM_COLUMNS,
            child_doctype='Sales Invoice Item',
            prepend_empty=False  # rows already include the leading empty column
        )

    # Pour CSV et XLSX : format tabulaire simple
    flat_cols = [
        {'label': 'Numéro de facture',   'name': 'name'},
        {'label': 'Client',              'name': 'customer'},
        {'label': 'Date',                'name': 'posting_date'},
        {'label': 'Date d\'échéance',    'name': 'due_date'},
        {'label': 'Statut',              'name': 'status'},
        {'label': 'Type',                'name': 'invoice_type'},
        {'label': 'Mode de paiement',    'name': 'payment_method'},
        {'label': 'Sous-total (XAF)',    'name': 'subtotal'},
        {'label': 'Taxe (XAF)',          'name': 'tax_amount'},
        {'label': 'Total (XAF)',         'name': 'total_amount'},
        {'label': 'Code article',        'name': 'item_code'},
        {'label': 'Nom article',         'name': 'item_name'},
        {'label': 'Quantité',            'name': 'qty'},
        {'label': 'Prix unitaire (XAF)', 'name': 'rate'},
        {'label': 'Total ligne (XAF)',   'name': 'amount'},
    ]

    rows = []
    for inv in invoices:
        client_name = inv.client.name if inv.client else ''
        inv_date = inv.created_at.date().strftime('%d/%m/%Y')
        due_date = inv.due_date.strftime('%d/%m/%Y') if inv.due_date else ''
        items = list(inv.items.all())

        if not items:
            rows.append([
                inv.invoice_number, client_name, inv_date, due_date,
                inv.status, inv.invoice_type, inv.payment_method or '',
                float(inv.subtotal), float(inv.tax_amount), float(inv.total_amount),
                '', '', '', '', ''
            ])
        else:
            for i, item in enumerate(items):
                item_code = (item.product.reference or item.product.name[:20]) if item.product else ''
                item_name = item.product.name if item.product else (item.description or '')
                if i == 0:
                    rows.append([
                        inv.invoice_number, client_name, inv_date, due_date,
                        inv.status, inv.invoice_type, inv.payment_method or '',
                        float(inv.subtotal), float(inv.tax_amount), float(inv.total_amount),
                        item_code, item_name,
                        float(item.quantity), float(item.unit_price), float(item.total_price)
                    ])
                else:
                    rows.append([
                        '', '', '', '', '', '', '', '', '', '',
                        item_code, item_name,
                        float(item.quantity), float(item.unit_price), float(item.total_price)
                    ])

    return _dispatch_response(fmt, f'factures_{date_from}_{date_to}',
                               flat_cols, rows,
                               doctype='Sales Invoice',
                               sheet_title='Factures')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_payments(request):
    """Export paiements reçus → ERPNext Payment Entry"""
    if not _require_admin(request.user):
        return Response({'detail': 'Accès réservé aux administrateurs.'}, status=drf_status.HTTP_403_FORBIDDEN)

    from apps.invoicing.models import Invoice

    org = _get_org(request.user)
    if not org:
        return Response({'detail': 'Organisation introuvable.'}, status=drf_status.HTTP_400_BAD_REQUEST)

    fmt = request.GET.get('format', 'csv')
    date_from = _parse_date(request.GET.get('date_from'), date(date.today().year, 1, 1))
    date_to = _parse_date(request.GET.get('date_to'), date.today())

    invoices = Invoice.objects.filter(
        organization=org,
        status='paid',
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).select_related('client').order_by('created_at')

    company = org.name

    rows = []
    for inv in invoices:
        client_name = inv.client.name if inv.client else ''
        pay_date = inv.updated_at.date().strftime('%d/%m/%Y')
        mode = _map_payment_mode(inv.payment_method)
        rows.append([
            '',                         # ID (vide = nouveau)
            'Receive',                  # payment_type
            pay_date,                   # posting_date
            company,                    # company
            'Customer',                 # party_type
            client_name,                # party
            float(inv.total_amount),    # paid_amount
            'XAF',                      # paid_from_account_currency
            mode,                       # mode_of_payment
            inv.invoice_number,         # reference_no
            pay_date,                   # reference_date
            f'Paiement {inv.invoice_number}',  # remarks
        ])

    return _dispatch_response(fmt, f'paiements_{date_from}_{date_to}',
                               PAYMENT_COLUMNS, rows,
                               doctype='Payment Entry',
                               sheet_title='Paiements')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_purchase_orders(request):
    """Export bons de commande → ERPNext Purchase Order"""
    if not _require_admin(request.user):
        return Response({'detail': 'Accès réservé aux administrateurs.'}, status=drf_status.HTTP_403_FORBIDDEN)

    org = _get_org(request.user)
    if not org:
        return Response({'detail': 'Organisation introuvable.'}, status=drf_status.HTTP_400_BAD_REQUEST)

    fmt = request.GET.get('format', 'csv')
    date_from = _parse_date(request.GET.get('date_from'), date(date.today().year, 1, 1))
    date_to = _parse_date(request.GET.get('date_to'), date.today())

    try:
        from apps.purchase_orders.models import PurchaseOrder
        orders = PurchaseOrder.objects.filter(
            organization=org,
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).select_related('supplier').prefetch_related('items__product').order_by('created_at')
    except Exception:
        orders = []

    company = org.name

    flat_cols = [
        {'label': 'Numéro BC',         'name': 'name'},
        {'label': 'Fournisseur',       'name': 'supplier'},
        {'label': 'Date',              'name': 'transaction_date'},
        {'label': 'Société',           'name': 'company'},
        {'label': 'Statut',            'name': 'status'},
        {'label': 'Devise',            'name': 'currency'},
        {'label': 'Code article',      'name': 'item_code'},
        {'label': 'Nom article',       'name': 'item_name'},
        {'label': 'Quantité',          'name': 'qty'},
        {'label': 'Prix unitaire (XAF)','name': 'rate'},
        {'label': 'Total ligne (XAF)', 'name': 'amount'},
        {'label': 'Date livraison',    'name': 'schedule_date'},
    ]

    rows = []
    for order in orders:
        items = list(order.items.all())
        supplier_name = order.supplier.name if order.supplier else ''
        order_date = order.created_at.date().strftime('%d/%m/%Y')
        po_number = getattr(order, 'po_number', None) or getattr(order, 'name', str(order.id)[:10])
        po_status = getattr(order, 'status', '')

        if not items:
            rows.append([po_number, supplier_name, order_date, company, po_status, 'XAF', '', '', '', '', '', ''])
        else:
            for i, item in enumerate(items):
                item_code = (item.product.reference or item.product.name[:20]) if item.product else ''
                item_name = item.product.name if item.product else getattr(item, 'description', '')
                sched = getattr(item, 'delivery_date', None) or getattr(item, 'schedule_date', None)
                sched_str = sched.strftime('%d/%m/%Y') if sched else ''
                rows.append([
                    po_number if i == 0 else '',
                    supplier_name if i == 0 else '',
                    order_date if i == 0 else '',
                    company if i == 0 else '',
                    po_status if i == 0 else '',
                    'XAF' if i == 0 else '',
                    item_code, item_name,
                    float(item.quantity) if hasattr(item, 'quantity') else 0,
                    float(item.unit_price) if hasattr(item, 'unit_price') else 0,
                    float(item.total_price) if hasattr(item, 'total_price') else 0,
                    sched_str,
                ])

    return _dispatch_response(fmt, f'bons_commande_{date_from}_{date_to}',
                               flat_cols, rows,
                               doctype='Purchase Order',
                               sheet_title='Bons de commande')


# ──────────────────────────────────────────────
#  Utilitaires de mapping
# ──────────────────────────────────────────────

UOM_MAP = {
    'piece': 'Nos',
    'box': 'Box',
    'tablet': 'Tab',
    'capsule': 'Cap',
    'blister': 'Bls',
    'bottle': 'Bottle',
    'vial': 'Amp',
    'sachet': 'Sachet',
    'tube': 'Tube',
    'ml': 'Ml',
    'l': 'Liter',
    'g': 'Gram',
    'kg': 'Kg',
}


def _map_uom(unit):
    return UOM_MAP.get(unit, 'Nos')


def _map_payment_mode(method):
    mapping = {
        'mobile_money': 'Mobile Money',
        'cash': 'Cash',
        '': 'Cash',
    }
    return mapping.get(method or '', 'Cash')
