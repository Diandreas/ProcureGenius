"""
Service d'export des statistiques du dashboard en PDF et Excel
"""
from io import BytesIO
from datetime import datetime
from django.utils import timezone
from typing import Dict
import logging

logger = logging.getLogger(__name__)


class DashboardExportService:
    """Service pour exporter les statistiques du dashboard"""

    def __init__(self, stats_data: Dict, user):
        """
        Initialize export service

        Args:
            stats_data: Dictionnaire des statistiques à exporter
            user: Utilisateur Django
        """
        self.stats_data = stats_data
        self.user = user
        self.metadata = stats_data.get('metadata', {})

    def export_to_pdf(self) -> BytesIO:
        """
        Exporte les statistiques en PDF avec graphiques

        Returns:
            BytesIO: Buffer contenant le PDF
        """
        try:
            from weasyprint import HTML, CSS
            from django.template.loader import render_to_string
            from django.conf import settings

            # Préparer le contexte pour le template
            context = {
                'stats_data': self.stats_data,
                'metadata': self.metadata,
                'user_name': self.user.get_full_name() or self.user.username,
                'current_datetime': datetime.now().strftime('%d/%m/%Y à %H:%M'),
                'generated_date': datetime.now().strftime('%d/%m/%Y'),
            }

            # Rendu HTML
            html_string = render_to_string('analytics/dashboard_report.html', context)

            # Générer le PDF avec WeasyPrint
            html = HTML(string=html_string, base_url=settings.BASE_DIR)
            pdf_bytes = html.write_pdf()

            # Convertir en BytesIO pour compatibilité avec l'API existante
            buffer = BytesIO(pdf_bytes)
            buffer.seek(0)

            return buffer

        except Exception as e:
            logger.error(f"Error generating PDF with WeasyPrint: {e}")
            raise

    def export_to_excel(self) -> BytesIO:
        """
        Exporte les statistiques en Excel avec plusieurs feuilles

        Returns:
            BytesIO: Buffer contenant le fichier Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            workbook = openpyxl.Workbook()

            # Supprimer la feuille par défaut
            workbook.remove(workbook.active)

            # Style pour les en-têtes
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Feuille Résumé
            ws_summary = workbook.create_sheet("Résumé")
            ws_summary.append(["Tableau de Bord - Résumé"])
            ws_summary.merge_cells('A1:B1')
            ws_summary['A1'].font = Font(bold=True, size=16, color="1976D2")
            ws_summary.append([])

            ws_summary.append(["Période", f"{self.metadata.get('start_date')} - {self.metadata.get('end_date')}"])
            ws_summary.append(["Nombre de jours", self.metadata.get('period_days', 'N/A')])
            ws_summary.append(["Généré le", datetime.now().strftime('%d/%m/%Y %H:%M')])
            ws_summary.append([])

            # Données financières
            if 'financial' in self.stats_data:
                ws_summary.append(["Résumé Financier"])
                ws_summary.append(["Indicateur", "Valeur"])
                financial = self.stats_data['financial']
                ws_summary.append(["Revenus", financial.get('revenue', 0)])
                ws_summary.append(["Dépenses", financial.get('expenses', 0)])
                ws_summary.append(["Profit Net", financial.get('net_profit', 0)])
                ws_summary.append(["Marge bénéficiaire", f"{financial.get('profit_margin', 0):.2f}%"])

            # Feuille Factures
            if 'invoices' in self.stats_data:
                ws_invoices = workbook.create_sheet("Factures")
                ws_invoices.append(["Statistiques des Factures"])
                ws_invoices.merge_cells('A1:B1')
                ws_invoices['A1'].font = Font(bold=True, size=14)
                ws_invoices.append([])

                invoices = self.stats_data['invoices']
                ws_invoices.append(["Métrique", "Valeur"])
                ws_invoices.append(["Total factures", invoices.get('total', 0)])

                by_status = invoices.get('by_status', {})
                for status, count in by_status.items():
                    ws_invoices.append([f"Statut: {status}", count])

                if 'period' in invoices:
                    ws_invoices.append([])
                    ws_invoices.append(["Période analysée"])
                    period = invoices['period']
                    ws_invoices.append(["Nouvelles factures", period.get('count', 0)])
                    ws_invoices.append(["Montant total", period.get('total_amount', 0)])
                    ws_invoices.append(["Montant payé", period.get('paid_amount', 0)])
                    ws_invoices.append(["Taux de paiement", f"{period.get('payment_rate', 0):.2f}%"])

            # Feuille Bons de Commande
            if 'purchase_orders' in self.stats_data:
                ws_pos = workbook.create_sheet("Bons de Commande")
                ws_pos.append(["Statistiques des Bons de Commande"])
                ws_pos.merge_cells('A1:B1')
                ws_pos['A1'].font = Font(bold=True, size=14)
                ws_pos.append([])

                pos = self.stats_data['purchase_orders']
                ws_pos.append(["Métrique", "Valeur"])
                ws_pos.append(["Total BCs", pos.get('total', 0)])

                by_status = pos.get('by_status', {})
                for status, count in by_status.items():
                    ws_pos.append([f"Statut: {status}", count])

                if 'period' in pos:
                    ws_pos.append([])
                    ws_pos.append(["Période analysée"])
                    period = pos['period']
                    ws_pos.append(["Nouveaux BCs", period.get('count', 0)])
                    ws_pos.append(["Montant total", period.get('total_amount', 0)])

            # Auto-ajuster les colonnes
            for sheet in workbook.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return buffer

        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            raise

    def _format_comparison(self, data: Dict, key: str) -> str:
        """Formate la comparaison avec période précédente"""
        if 'comparison' not in data:
            return ''

        comparison = data['comparison']
        change_key = f"{key}_change" if f"{key}_change" in comparison else 'change'
        percent_key = f"{key}_percent_change" if f"{key}_percent_change" in comparison else 'percent_change'

        change = comparison.get(change_key, 0)
        percent = comparison.get(percent_key, 0)

        if change == 0:
            return '='

        symbol = '▲' if change > 0 else '▼'
        return f"{symbol} {abs(percent):.1f}%"
