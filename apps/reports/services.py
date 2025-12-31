"""
Services pour la génération de rapports PDF et Excel
"""
from datetime import datetime
from io import BytesIO
from django.core.files.base import ContentFile
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from .models import Report, ReportTemplate


class BaseReportService:
    """Service de base pour la génération de rapports"""

    def __init__(self, report_type, user=None):
        self.report_type = report_type
        self.user = user

    def create_report_record(self, format='pdf', parameters=None, template=None):
        """Créer un enregistrement de rapport"""
        report = Report.objects.create(
            report_type=self.report_type,
            format=format,
            parameters=parameters or {},
            template=template,
            generated_by=self.user,
            status='processing'
        )
        return report

    def mark_report_completed(self, report, file_content, filename):
        """Marquer un rapport comme terminé"""
        report.file_path.save(filename, ContentFile(file_content), save=False)
        report.file_size = len(file_content)
        report.status = 'completed'
        report.completed_at = timezone.now()
        report.save()
        return report

    def mark_report_failed(self, report, error_message):
        """Marquer un rapport comme échoué"""
        report.status = 'failed'
        report.error_message = str(error_message)
        report.completed_at = timezone.now()
        report.save()
        return report


class SupplierReportService(BaseReportService):
    """Service de génération de rapports pour les fournisseurs"""

    def __init__(self, user=None):
        super().__init__('supplier', user)

    def generate(self, supplier_id, format='pdf', date_start=None, date_end=None):
        """Générer un rapport pour un fournisseur spécifique"""
        from apps.suppliers.models import Supplier
        from apps.purchase_orders.models import PurchaseOrder

        parameters = {
            'supplier_id': str(supplier_id),
            'date_start': date_start.isoformat() if date_start else None,
            'date_end': date_end.isoformat() if date_end else None,
        }

        report = self.create_report_record(format=format, parameters=parameters)

        try:
            supplier = Supplier.objects.get(id=supplier_id)

            # Collecter les données
            pos = PurchaseOrder.objects.filter(supplier=supplier)
            if date_start:
                pos = pos.filter(created_at__gte=date_start)
            if date_end:
                pos = pos.filter(created_at__lte=date_end)

            data = {
                'supplier': supplier,
                'total_orders': pos.count(),
                'total_amount': pos.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
                'avg_order_value': pos.aggregate(Avg('total_amount'))['total_amount__avg'] or 0,
                'orders': pos.order_by('-created_at')[:20],
                'date_start': date_start,
                'date_end': date_end,
                'generated_at': timezone.now(),
            }

            # Générer selon le format
            if format == 'pdf':
                file_content = self._generate_pdf(data)
                filename = f"rapport_fournisseur_{supplier_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            elif format == 'xlsx':
                file_content = self._generate_excel(data)
                filename = f"rapport_fournisseur_{supplier_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            else:
                raise ValueError(f"Format non supporté: {format}")

            self.mark_report_completed(report, file_content, filename)
            return report

        except Exception as e:
            self.mark_report_failed(report, str(e))
            raise

    def _generate_pdf(self, data):
        """Générer un PDF"""
        try:
            from weasyprint import HTML, CSS
            from django.template.loader import render_to_string
            from django.conf import settings

            # Préparer le contexte pour le template
            context = {
                'supplier': data['supplier'],
                'total_orders': data['total_orders'],
                'total_amount': data['total_amount'],
                'avg_order_value': data['avg_order_value'],
                'orders': data['orders'],
                'date_start': data['date_start'],
                'date_end': data['date_end'],
                'generated_at': data['generated_at'],
                # Ajouter les données nécessaires pour le template existant
                'total_spent': data['total_amount'],
                'po_by_status': [],  # On peut ajouter cette logique si nécessaire
                'top_products': [],  # On peut ajouter cette logique si nécessaire
                'recent_activity': data['orders'][:20],  # Utiliser les commandes comme activité récente
                'organization': getattr(data['supplier'], 'organization', None),  # Si disponible
            }

            # Rendu HTML
            html_string = render_to_string('reports/pdf/supplier_report.html', context)

            # Générer le PDF avec WeasyPrint
            html = HTML(string=html_string, base_url=settings.BASE_DIR)
            pdf_bytes = html.write_pdf()

            return pdf_bytes

        except ImportError:
            # Fallback si weasyprint n'est pas installé
            return b"PDF generation requires weasyprint. Please install: pip install weasyprint"

    def _generate_excel(self, data):
        """Générer un fichier Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapport Fournisseur"

            # Header styling
            header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=14)

            # Title
            ws['A1'] = f"Rapport Fournisseur: {data['supplier'].name}"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')

            # Informations générales
            ws['A3'] = "Informations Générales"
            ws['A3'].font = header_font
            ws['A3'].fill = header_fill
            ws.merge_cells('A3:B3')

            info_rows = [
                ('Fournisseur', data['supplier'].name),
                ('Email', data['supplier'].email or 'N/A'),
                ('Téléphone', data['supplier'].phone or 'N/A'),
                ('Ville', data['supplier'].city or 'N/A'),
                ('Note', f"{data['supplier'].rating}/5" if data['supplier'].rating else 'N/A'),
            ]

            row = 4
            for label, value in info_rows:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

            # Statistiques
            ws[f'A{row+1}'] = "Statistiques"
            ws[f'A{row+1}'].font = header_font
            ws[f'A{row+1}'].fill = header_fill
            ws.merge_cells(f'A{row+1}:B{row+1}')

            stats_row = row + 2
            stats_rows = [
                ('Nombre de commandes', data['total_orders']),
                ('Montant total', f"{data['total_amount']:.2f} $"),
                ('Valeur moyenne par commande', f"{data['avg_order_value']:.2f} $"),
            ]

            for label, value in stats_rows:
                ws[f'A{stats_row}'] = label
                ws[f'B{stats_row}'] = value
                ws[f'A{stats_row}'].font = Font(bold=True)
                stats_row += 1

            # Dernières commandes
            if data['orders'].exists():
                orders_row = stats_row + 2
                ws[f'A{orders_row}'] = "Dernières Commandes"
                ws[f'A{orders_row}'].font = header_font
                ws[f'A{orders_row}'].fill = header_fill
                ws.merge_cells(f'A{orders_row}:D{orders_row}')

                # Headers
                orders_row += 1
                headers = ['N° BC', 'Date', 'Statut', 'Montant']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=orders_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Data
                orders_row += 1
                for order in data['orders']:
                    ws[f'A{orders_row}'] = order.po_number
                    ws[f'B{orders_row}'] = order.created_at.strftime('%Y-%m-%d')
                    ws[f'C{orders_row}'] = order.get_status_display()
                    ws[f'D{orders_row}'] = f"{order.total_amount:.2f} $"
                    orders_row += 1

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            excel_content = buffer.getvalue()
            buffer.close()
            return excel_content

        except ImportError:
            # Fallback si openpyxl n'est pas installé
            return b"Excel generation requires openpyxl. Please install: pip install openpyxl"


# TODO: Créer les autres services de rapports
class ProductReportService(BaseReportService):
    """Service de génération de rapports pour les produits"""
    pass


class InvoiceReportService(BaseReportService):
    """Service de génération de rapports pour les factures"""
    pass


class PurchaseOrderReportService(BaseReportService):
    """Service de génération de rapports pour les bons de commande"""
    pass
