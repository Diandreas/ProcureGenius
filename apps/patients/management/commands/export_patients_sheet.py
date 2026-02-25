"""
Commande pour exporter la liste des patients (numéro + nom + téléphone)
en fichier Excel et l'envoyer par email.

Usage:
    python manage.py export_patients_sheet
    python manage.py export_patients_sheet --email direction@centrejulianna.com
    python manage.py export_patients_sheet --dry-run
"""
import io
import smtplib
from datetime import date
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.core.management.base import BaseCommand, CommandError


GENDER_LABELS = {'M': 'Masculin', 'F': 'Féminin', '': '—'}
BLOOD_TYPE_LABELS = {
    'A+': 'A+', 'A-': 'A−', 'B+': 'B+', 'B-': 'B−',
    'AB+': 'AB+', 'AB-': 'AB−', 'O+': 'O+', 'O-': 'O−', '': '—',
}


class Command(BaseCommand):
    help = "Exporte la liste complète des patients en Excel et l'envoie par email"

    def add_arguments(self, parser):
        parser.add_argument(
            '--email',
            type=str,
            default=None,
            help='Adresse email destinataire (défaut: admins actifs de l\'organisation)'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Génère le fichier Excel sans envoyer l\'email'
        )

    def handle(self, *args, **options):
        from apps.accounts.models import Organization, EmailConfiguration, Client

        # --- Organisation ---
        org = Organization.objects.first()
        if not org:
            raise CommandError("Aucune organisation trouvée.")
        self.stdout.write(f"Organisation : {org.name}")

        # --- Récupérer tous les patients ---
        patients = Client.objects.filter(
            organization=org,
            client_type__in=['patient', 'both'],
        ).order_by('patient_number')

        total = patients.count()
        self.stdout.write(f"Patients trouvés : {total}")

        if total == 0:
            self.stdout.write(self.style.WARNING("Aucun patient à exporter."))
            return

        # --- Générer le fichier Excel ---
        excel_buffer = self._build_excel(patients, org)
        filename = f"patients_{org.name.replace(' ', '_')}_{date.today().isoformat()}.xlsx"
        self.stdout.write(f"Fichier Excel généré : {filename} ({total} patients)")

        if options['dry_run']:
            self.stdout.write(self.style.WARNING("Mode dry-run : email non envoyé."))
            # Sauvegarde locale pour vérification
            with open(filename, 'wb') as f:
                f.write(excel_buffer.getvalue())
            self.stdout.write(f"Fichier sauvegardé localement : {filename}")
            return

        # --- Destinataires ---
        recipients = self._get_recipients(org, options['email'])
        if not recipients:
            raise CommandError(
                "Aucun destinataire trouvé. Utilisez --email pour spécifier une adresse."
            )
        self.stdout.write(f"Destinataire(s) : {', '.join(recipients)}")

        # --- Envoi email ---
        email_config = EmailConfiguration.objects.filter(organization=org).first()
        if not email_config:
            raise CommandError(
                "Aucune configuration SMTP trouvée. Lancez d'abord : python manage.py configure_smtp"
            )

        subject = f"[{org.name}] Liste des patients — {date.today().strftime('%d/%m/%Y')}"
        html_body = self._build_email_body(org, total)

        self._send_email(email_config, recipients, subject, html_body, excel_buffer, filename)
        self.stdout.write(
            self.style.SUCCESS(f"\nFichier envoyé à {', '.join(recipients)} !")
        )

    # ------------------------------------------------------------------
    # Construction du fichier Excel
    # ------------------------------------------------------------------

    def _build_excel(self, patients, org):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patients"

        # Couleurs
        HEADER_FILL = PatternFill("solid", fgColor="1A5276")
        ALT_FILL    = PatternFill("solid", fgColor="EBF5FB")
        THIN_BORDER = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA'),
        )

        # Titre du document
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Liste des patients — {org.name} — {date.today().strftime('%d/%m/%Y')}"
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        title_cell.fill = PatternFill("solid", fgColor="0D3B66")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # En-têtes
        headers = [
            "N° Patient",
            "Nom complet",
            "Téléphone",
            "Email",
            "Date de naissance",
            "Genre",
            "Groupe sanguin",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[2].height = 22

        # Données patients
        for row_idx, patient in enumerate(patients, start=3):
            is_even = (row_idx % 2 == 0)
            row_fill = ALT_FILL if is_even else PatternFill("solid", fgColor="FFFFFF")

            dob = patient.date_of_birth.strftime('%d/%m/%Y') if patient.date_of_birth else '—'
            gender = GENDER_LABELS.get(patient.gender or '', '—')
            blood = BLOOD_TYPE_LABELS.get(patient.blood_type or '', '—')

            values = [
                patient.patient_number or '—',
                patient.name or '—',
                patient.phone or '—',
                patient.email or '—',
                dob,
                gender,
                blood,
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Calibri', size=10)
                cell.fill = row_fill
                cell.alignment = Alignment(
                    horizontal='center' if col_idx != 2 else 'left',
                    vertical='center'
                )
                cell.border = THIN_BORDER
            ws.row_dimensions[row_idx].height = 18

        # Largeurs des colonnes
        col_widths = [14, 30, 18, 28, 16, 12, 14]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Figer les 2 premières lignes
        ws.freeze_panes = 'A3'

        # Filtre automatique sur les en-têtes
        ws.auto_filter.ref = f"A2:G{len(list(patients)) + 2}"

        # Pied de page (ligne de total)
        footer_row = len(list(patients)) + 3
        ws.merge_cells(f'A{footer_row}:F{footer_row}')
        footer_cell = ws[f'A{footer_row}']
        footer_cell.value = f"Total : {patients.count()} patient(s) enregistré(s)"
        footer_cell.font = Font(name='Calibri', bold=True, size=10, color='0D3B66')
        footer_cell.alignment = Alignment(horizontal='right', vertical='center')
        ws.row_dimensions[footer_row].height = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ------------------------------------------------------------------
    # Corps de l'email HTML
    # ------------------------------------------------------------------

    def _build_email_body(self, org, total):
        today = date.today().strftime('%d/%m/%Y')
        return f"""
        <html>
        <body style="font-family: Calibri, Arial, sans-serif; color: #222; background: #f5f7fa;">
          <div style="max-width:600px; margin:30px auto; background:#fff;
                      border-radius:8px; overflow:hidden;
                      box-shadow:0 2px 8px rgba(0,0,0,.12);">

            <div style="background:#0D3B66; padding:24px 32px;">
              <h2 style="color:#fff; margin:0; font-size:20px;">
                Liste des patients — {org.name}
              </h2>
              <p style="color:#BDD7F0; margin:6px 0 0; font-size:13px;">
                Exportée le {today}
              </p>
            </div>

            <div style="padding:28px 32px;">
              <p style="margin:0 0 16px;">Bonjour,</p>
              <p>Veuillez trouver en pièce jointe le fichier Excel contenant
                 la liste complète des patients enregistrés dans le système.</p>

              <div style="background:#EBF5FB; border-left:4px solid #1A5276;
                          padding:14px 18px; border-radius:4px; margin:20px 0;">
                <strong style="font-size:15px; color:#0D3B66;">
                  {total} patient(s) enregistré(s)
                </strong>
              </div>

              <p>Le fichier contient les informations suivantes :</p>
              <ul style="padding-left:20px; line-height:1.8;">
                <li>Numéro de patient</li>
                <li>Nom complet</li>
                <li>Téléphone</li>
                <li>Email</li>
                <li>Date de naissance</li>
                <li>Genre</li>
                <li>Groupe sanguin</li>
              </ul>

              <p style="margin-top:24px; color:#666; font-size:12px;">
                Ce rapport est généré automatiquement par le système ProcureGenius.<br>
                Ne pas répondre à cet email.
              </p>
            </div>

            <div style="background:#f0f4f8; padding:14px 32px; text-align:center;
                        font-size:11px; color:#888;">
              {org.name} &mdash; ProcureGenius
            </div>
          </div>
        </body>
        </html>
        """

    # ------------------------------------------------------------------
    # Envoi email avec pièce jointe
    # ------------------------------------------------------------------

    def _get_recipients(self, org, email_override):
        if email_override:
            return [email_override]
        recipients = list(
            org.users.filter(
                is_active=True,
                role__in=['admin', 'manager']
            ).exclude(email='').values_list('email', flat=True)
        )
        if not recipients:
            recipients = list(
                org.users.filter(is_active=True).exclude(email='').values_list('email', flat=True)[:5]
            )
        return recipients

    def _send_email(self, email_config, recipients, subject, html_body, excel_buffer, filename):
        password = email_config.get_decrypted_password()
        if not password:
            raise CommandError("Impossible de déchiffrer le mot de passe SMTP.")

        from_header = f"{email_config.default_from_name} <{email_config.default_from_email}>"

        msg = MIMEMultipart('mixed')
        msg['Subject'] = subject
        msg['From'] = from_header
        msg['To'] = ', '.join(recipients)

        # Partie HTML
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        # Pièce jointe Excel
        attachment = MIMEApplication(
            excel_buffer.read(),
            Name=filename
        )
        attachment['Content-Disposition'] = f'attachment; filename="{filename}"'
        msg.attach(attachment)

        try:
            if email_config.use_ssl:
                server = smtplib.SMTP_SSL(
                    email_config.smtp_host,
                    email_config.smtp_port,
                    timeout=30
                )
            else:
                server = smtplib.SMTP(
                    email_config.smtp_host,
                    email_config.smtp_port,
                    timeout=30
                )
                if email_config.use_tls:
                    server.starttls()

            server.login(email_config.smtp_username, password)
            server.sendmail(email_config.default_from_email, recipients, msg.as_string())
            server.quit()

        except smtplib.SMTPException as e:
            raise CommandError(f"Erreur lors de l'envoi de l'email : {e}")
